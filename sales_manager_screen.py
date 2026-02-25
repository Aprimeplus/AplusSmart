import tkinter as tk
from tkinter import ttk, filedialog
from customtkinter import (CTkFrame, CTkLabel, CTkFont, CTkButton,
                               CTkScrollableFrame, CTkInputDialog, CTkToplevel, CTkEntry,
                               CTkOptionMenu, CTkRadioButton, CTkTabview, CTkCheckBox)
from tkinter import messagebox
import pandas as pd
from datetime import datetime
import psycopg2.errors
import psycopg2.extras
import traceback
import utils
import os

# เพิ่ม matplotlib สำหรับกราฟ
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib
matplotlib.use('TkAgg')

# --- นำเข้า Class ที่จำเป็น ---
from history_windows import SOPopupWindow
from daily_report_widget import DailyReportWidget

class ToastNotification(CTkToplevel):
    """หน้าต่างแจ้งเตือนมุมขวาล่าง เด้งโชว์แล้วหายไปเอง (ไม่บล็อกการทำงาน)"""
    def __init__(self, master, title, message, duration=10000, color="#16A34A"):
        super().__init__(master)
        
        # ลบกรอบหน้าต่างออก (ไร้ขอบ)
        self.overrideredirect(True)
        self.attributes("-topmost", True) # ให้อยู่บนสุดเสมอ
        
        # ตั้งค่าสีและ Layout
        self.configure(fg_color=color)
        
        # จัดข้อความ
        CTkLabel(self, text=title, font=CTkFont(size=14, weight="bold"), text_color="white").pack(padx=20, pady=(10, 2), anchor="w")
        CTkLabel(self, text=message, font=CTkFont(size=12), text_color="white", justify="left").pack(padx=20, pady=(0, 10), anchor="w")

        self.update_idletasks()
        width = self.winfo_reqwidth()
        height = self.winfo_reqheight()

        # คำนวณตำแหน่งมุมขวาล่างของหน้าจอ
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        # วางที่มุมขวาล่าง (ห่างขอบนิดหน่อย)
        x = screen_width - width - 30
        y = screen_height - height - 60 
        
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # ทำให้คลิกที่แจ้งเตือนแล้วหายไปทันที (เผื่อไม่อยากรอ 10 วิ)
        self.bind("<Button-1>", lambda e: self.destroy())
        for child in self.winfo_children():
            child.bind("<Button-1>", lambda e: self.destroy())

        # ตั้งเวลาทำลายตัวเอง (10 วินาที = 10000 ms)
        self.after(duration, self.destroy)

class SalesManagerScreen(CTkFrame):
    def __init__(self, master, app_container, user_key=None, user_name=None, user_role=None):
        super().__init__(master)
        self.app_container = app_container
        self.user_key = user_key
        self.user_name = user_name
        self.user_role = user_role
        
        self.label_font = CTkFont(size=14, weight="bold")
        self.entry_font = CTkFont(size=14)
        
        # --- เตรียมตัวแปรสำหรับ SOPopupWindow ---
        self.so_popup = None
        self._so_create_string_vars()
        self.sale_theme = self.app_container.THEME.get("sale", {"bg": "white", "primary": "#3B82F6"})
        
        self.pg_engine = self.app_container.pg_engine
        
        # ตัวแปรสำหรับ Filter กราฟ
        self._init_chart_filters()
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) 

        # --- 1. Header ---
        self._create_header()

        # --- 2. TabView ---
        self.tab_view = CTkTabview(self, corner_radius=10, border_width=1)
        self.tab_view.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")

        # ✅ เพิ่มแท็บ 1: รายการรออนุมัติ (เป็นหน้าแรก)
        self.approval_tab = self.tab_view.add("🗳️ รายการรออนุมัติ (SM Approval)")
        
        # แท็บเดิม
        self.daily_report_tab = self.tab_view.add("📅 รายงานประจำวัน (SO Report)")
        self.master_tab = self.tab_view.add("🛠️ ค้นหาและจัดการ (Master)")

        # สร้างเนื้อหาในแต่ละ Tab
        self._create_approval_tab(self.approval_tab)
        self._create_daily_report_widget(self.daily_report_tab) 
        self._create_master_tab(self.master_tab)            
        
        # ตั้งค่าหน้าแรกที่เปิดขึ้นมา
        self.tab_view.set("🗳️ รายการรออนุมัติ (SM Approval)")

        self.known_pending_so_ids = set() # เก็บ ID ของ SO ที่เคยแจ้งเตือนไปแล้ว
        self.reminder_timer_count = 0     # ตัวนับเวลาสำหรับเตือนทุก 10 นาที
        self.noti_job_id = None
        
        # เริ่มทำงานระบบ Noti เบื้องหลัง (หน่วงเวลา 3 วินาทีหลังเปิดหน้าจอ)
        self.after(3000, self._start_notification_system)
        
        # ดักจับตอนปิดหน้าจอให้หยุด Noti ด้วย
        self.bind("<Destroy>", self._on_destroy)
    
    def _start_notification_system(self):
        """ลูปตรวจสอบ Noti ทุกๆ 1 นาที"""
        self._check_pending_approvals()
        
        # ลูปเรียกตัวเองทุกๆ 60 วินาที (60,000 ms)
        self.noti_job_id = self.after(60000, self._start_notification_system)

    def _check_pending_approvals(self):
        """เช็ค DB ว่ามี SO ใหม่ หรือต้องทวงงานหรือไม่"""
        try:
            query = """
                SELECT id, so_number, customer_name 
                FROM commissions 
                WHERE status = 'Pending Sale Manager Approval' AND is_active = 1
            """
            df = pd.read_sql_query(query, self.pg_engine)
            
            current_pending_ids = set(df['id'].tolist())
            pending_count = len(current_pending_ids)
            
            # 1. เช็คว่ามี SO "ใหม่" เข้ามาหรือไม่ (ID ที่เพิ่งโผล่มาใหม่)
            new_so_ids = current_pending_ids - self.known_pending_so_ids
            
            if new_so_ids:
                # มีงานใหม่เข้า! เด้งแจ้งเตือนแบบ 10 วินาที (สีเขียว/ฟ้า)
                new_count = len(new_so_ids)
                ToastNotification(
                    self.winfo_toplevel(), 
                    title="🔔 มี SO ใหม่รออนุมัติ!", 
                    message=f"มีคำขออนุมัติ SO เข้ามาใหม่จำนวน {new_count} รายการ\nกรุณาตรวจสอบในแท็บ 'รายการรออนุมัติ'",
                    duration=10000, 
                    color="#2563EB" # สีฟ้า
                )
                # อัปเดตรายการที่รู้จักแล้ว และรีเฟรชตารางโชว์งานใหม่
                self.known_pending_so_ids = current_pending_ids
                
                # --- ย้ายบรรทัดนี้ไปไว้ใน after() เพื่อไม่ให้ขัดจังหวะ ---
                self.after(100, self._load_approval_data)
                
            # 2. ระบบทวงงาน (Reminder) ถ้ายังมีงานค้าง
            elif pending_count > 0:
                self.reminder_timer_count += 1
                
                # เช็คทุก 1 นาที ดังนั้น 10 รอบ = 10 นาที
                if self.reminder_timer_count >= 10: 
                    ToastNotification(
                        self.winfo_toplevel(), 
                        title="⚠️ แจ้งเตือนงานค้าง!", 
                        message=f"คุณยังมี SO รอการอนุมัติค้างอยู่ {pending_count} รายการ",
                        duration=10000, 
                        color="#EA580C" # สีส้ม
                    )
                    self.reminder_timer_count = 0 # รีเซ็ตตัวนับ
            else:
                # ไม่มีงานค้างเลย รีเซ็ตเวลา
                self.reminder_timer_count = 0
                self.known_pending_so_ids.clear()
                
        except Exception as e:
            print(f"Noti System Error: {e}")

    def _on_destroy(self, event):
        """หยุด Loop เมื่อปิดหน้าจอ"""
        if hasattr(event, 'widget') and event.widget is self:
            if self.noti_job_id:
                self.after_cancel(self.noti_job_id)

    def _init_chart_filters(self):
        """สร้างตัวแปรสำหรับ Filter กราฟ"""
        now = datetime.now()
        self.thai_months = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", 
                           "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
        self.chart_month_var = tk.StringVar(value=self.thai_months[now.month - 1])
        self.chart_year_var = tk.StringVar(value=str(now.year + 543))

    def _create_header(self):
        header_frame = CTkFrame(self, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(10,0))
        
        CTkLabel(header_frame, text=f"Sale Manager Dashboard: {self.user_name}", font=CTkFont(size=22, weight="bold")).pack(side="left")
        
        button_frame = CTkFrame(header_frame, fg_color="transparent")
        button_frame.pack(side="right", padx=10)
        
        CTkButton(button_frame, text="🔄 Refresh All", command=self._refresh_all_tabs).pack(side="left", padx=5)
        
        CTkButton(button_frame, text="ออกจากระบบ", command=self.app_container.show_login_screen, 
                  fg_color="transparent", border_color="#D32F2F", 
                  text_color="#D32F2F", border_width=2, 
                  hover_color="#FFEBEE").pack(side="left", padx=5)

    def _refresh_all_tabs(self):
        """รีโหลดข้อมูลทุกแท็บ"""
        self._load_approval_data() # รีโหลดหน้าอนุมัติ
        self._update_rejection_chart() # รีโหลดกราฟ
        
        if hasattr(self, 'daily_report_widget'):
            self.daily_report_widget.load_report_data()
            if hasattr(self.daily_report_widget, 'dashboard_view'):
                 self.daily_report_widget.dashboard_view._update_chart()
        
        self._load_master_data() # รีโหลด Master Tab

    # =========================================================================
    # ✅ NEW TAB: SM APPROVAL พร้อม Dashboard กราฟ
    # =========================================================================
    def _create_approval_tab(self, parent):
        """สร้างแท็บรออนุมัติ แบ่งเป็น Dashboard 40% + List 60%"""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=4)  # Dashboard 40%
        parent.grid_rowconfigure(1, weight=6)  # List 60%

        # ========== ส่วนบน 40%: Dashboard กราฟ ==========
        dashboard_container = CTkFrame(parent, fg_color="#F3F4F6", corner_radius=10)
        dashboard_container.grid(row=0, column=0, padx=10, pady=(10, 5), sticky="nsew")
        dashboard_container.grid_columnconfigure(0, weight=1)
        dashboard_container.grid_rowconfigure(1, weight=1)

        # Header + Filters
        chart_header = CTkFrame(dashboard_container, fg_color="transparent")
        chart_header.grid(row=0, column=0, sticky="ew", padx=15, pady=10)
        
        CTkLabel(chart_header, text="📊 สถิติการตีกลับงานรายบุคคล", 
                font=CTkFont(size=16, weight="bold")).pack(side="left")

        filter_frame = CTkFrame(chart_header, fg_color="transparent")
        filter_frame.pack(side="right")
        
        # ✅ ปุ่ม Export Excel
        CTkButton(filter_frame, 
                 text="📥 Export Excel",
                 width=120,
                 height=32,
                 fg_color="#16A34A",
                 hover_color="#15803D",
                 font=CTkFont(size=12, weight="bold"),
                 command=self._export_rejection_to_excel).pack(side="left", padx=(0, 10))
        
        CTkLabel(filter_frame, text="รอบเดือน:", font=CTkFont(size=12)).pack(side="left", padx=5)
        CTkOptionMenu(filter_frame, variable=self.chart_month_var, values=self.thai_months, 
                      width=110, height=32,
                      command=lambda e: self._update_rejection_chart()).pack(side="left", padx=3)
        
        years = [str(y + 543) for y in range(datetime.now().year - 1, datetime.now().year + 2)]
        CTkOptionMenu(filter_frame, variable=self.chart_year_var, values=years, 
                      width=90, height=32,
                      command=lambda e: self._update_rejection_chart()).pack(side="left", padx=3)

        # พื้นที่กราฟ
        self.chart_area = CTkFrame(dashboard_container, fg_color="white", corner_radius=8)
        self.chart_area.grid(row=1, column=0, padx=15, pady=(5, 15), sticky="nsew")

        # ========== ส่วนล่าง 60%: รายการ SO ==========
        list_container = CTkFrame(parent, fg_color="transparent")
        list_container.grid(row=1, column=0, padx=10, pady=(5, 10), sticky="nsew")
        list_container.grid_columnconfigure(0, weight=1)
        list_container.grid_rowconfigure(1, weight=1)

        # Search Bar
        search_frame = CTkFrame(list_container, fg_color="#F3F4F6", corner_radius=8, height=60)
        search_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        search_frame.grid_propagate(False)
        
        CTkLabel(search_frame, text="🗳️ รายการรออนุมัติ", 
                font=CTkFont(size=14, weight="bold")).pack(side="left", padx=20)

        search_right = CTkFrame(search_frame, fg_color="transparent")
        search_right.pack(side="right", padx=20)
        
        self.approval_search_var = tk.StringVar()
        self.approval_search_entry = CTkEntry(
            search_right, 
            placeholder_text="🔍 ค้นหา SO หรือชื่อลูกค้า...", 
            width=300,
            height=36,
            textvariable=self.approval_search_var
        )
        self.approval_search_entry.pack(side="left", padx=(0, 8))
        self.approval_search_entry.bind("<Return>", lambda e: self._load_approval_data())

        CTkButton(search_right, text="ค้นหา", width=80, height=36,
                  command=self._load_approval_data).pack(side="left", padx=3)
        CTkButton(search_right, text="🔄 รีเฟรช", width=80, height=36, fg_color="gray",
                  command=self._load_approval_data).pack(side="left", padx=3)
        
        # รายการ SO
        self.approval_results_frame = CTkScrollableFrame(list_container, 
                                                         fg_color="white",
                                                         corner_radius=8)
        self.approval_results_frame.grid(row=1, column=0, sticky="nsew")
        self.approval_results_frame.grid_columnconfigure(0, weight=1)

        # โหลดข้อมูลครั้งแรก
        self.after(200, lambda: [self._update_rejection_chart(), self._load_approval_data()])

    def _update_rejection_chart(self):
        """วาดกราฟแท่งแนวนอนแสดงจำนวนการตีกลับของแต่ละ Sale"""
        # ล้างกราฟเก่า
        for widget in self.chart_area.winfo_children():
            widget.destroy()

        try:
            # ✅ ตั้งค่าฟอนต์ภาษาไทยสำหรับ Matplotlib
            try:
                plt.rcParams['font.family'] = 'TH Sarabun New'
            except:
                try:
                    plt.rcParams['font.family'] = 'Tahoma'
                except:
                    plt.rcParams['font.family'] = 'sans-serif'
            
            # ดึงข้อมูล
            month_idx = self.thai_months.index(self.chart_month_var.get()) + 1
            year_val = int(self.chart_year_var.get()) - 543

            # เอาบรรทัด AND u.manager_key ออกไป
            query = """
                SELECT u.sale_name, COUNT(c.id) as reject_count
                FROM commissions c
                JOIN sales_users u ON c.sale_key = u.sale_key
                WHERE c.status = 'Rejected by SM' 
                  AND c.commission_month = %s 
                  AND c.commission_year = %s
                  AND c.is_active = 1
                GROUP BY u.sale_name
                ORDER BY reject_count ASC
            """
            
            df = pd.read_sql_query(query, self.pg_engine, params=(month_idx, year_val))
            

            if df.empty:
                # ไม่มีข้อมูล - ✅ ตัวอักษรใหญ่ขึ้น
                empty_frame = CTkFrame(self.chart_area, fg_color="transparent")
                empty_frame.pack(expand=True, pady=30)
                
                CTkLabel(empty_frame, text="✅", 
                        font=CTkFont(size=48)).pack()
                CTkLabel(empty_frame, 
                        text="ไม่มีข้อมูลการตีกลับในรอบเดือนนี้",
                        font=CTkFont(size=15, weight="bold"),
                        text_color="#16A34A").pack(pady=(5, 0))
                return

            # วาดกราฟ - ✅ ปรับขนาดให้ใหญ่ขึ้น
            num_sales = len(df)
            fig_height = max(2.8, min(5.5, num_sales * 0.6))
            
            fig, ax = plt.subplots(figsize=(9.5, fig_height), dpi=100)
            
            # วาดแท่ง - ✅ แท่งหนาขึ้น
            bars = ax.barh(df['sale_name'], df['reject_count'], 
                          color='#FB923C', height=0.65)
            
            # ตั้งค่ากราฟ - ✅ เพิ่มขนาดตัวอักษร
            ax.set_title(f"จำนวนงานที่ถูกตีกลับ (รอบ {self.chart_month_var.get()} {self.chart_year_var.get()})",
                        fontweight='bold', fontsize=16, pad=15)
            ax.set_xlabel("จำนวนครั้ง", fontsize=14, fontweight='bold')
            ax.set_ylabel("เซลล์", fontsize=14, fontweight='bold')
            
            # ✅ ปรับขนาดตัวเลขบนแกน
            ax.tick_params(axis='both', which='major', labelsize=13)
            
            # ✅ บังคับให้แกน X เป็นจำนวนเต็ม
            ax.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
            
            # Grid - ✅ เส้น Grid หนาขึ้น
            ax.grid(axis='x', alpha=0.3, linestyle='--', linewidth=1.2)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_linewidth(1.5)
            ax.spines['bottom'].set_linewidth(1.5)

            # แสดงตัวเลขบนแท่ง - ✅ เพิ่มขนาดและความหนา
            for bar in bars:
                width = bar.get_width()
                if width > 0:
                    ax.text(width + 0.15, bar.get_y() + bar.get_height()/2,
                           f'{int(width)} ครั้ง',
                           va='center', 
                           fontsize=13,
                           fontweight='bold')

            plt.tight_layout(pad=1.5)
            
            # แสดงกราฟ
            canvas = FigureCanvasTkAgg(fig, master=self.chart_area)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
            
        except Exception as e:
            print(f"Chart Error: {traceback.format_exc()}")
            CTkLabel(self.chart_area, 
                    text=f"⚠️ เกิดข้อผิดพลาดในการโหลดกราฟ",
                    font=CTkFont(size=14, weight="bold"),
                    text_color="#DC2626").pack(expand=True, pady=20)

    def _export_rejection_to_excel(self):
        """Export ข้อมูลการตีกลับเป็น Excel พร้อมรายละเอียด"""
        try:
            month_idx = self.thai_months.index(self.chart_month_var.get()) + 1
            year_val = int(self.chart_year_var.get()) - 543
            
            # Query ข้อมูลรายละเอียดการตีกลับ
            query = """
                SELECT 
                    u.sale_name as "ชื่อเซลล์",
                    c.so_number as "เลขที่ SO",
                    c.customer_name as "ชื่อลูกค้า",
                    c.sales_service_amount as "ยอดขาย (บาท)",
                    c.rejection_reason as "เหตุผลที่ตีกลับ",
                    TO_CHAR(c.timestamp::timestamp, 'DD/MM/YYYY HH24:MI') as "วันที่ส่ง SO"
                FROM commissions c
                JOIN sales_users u ON c.sale_key = u.sale_key
                WHERE c.status = 'Rejected by SM' 
                  AND c.commission_month = %s 
                  AND c.commission_year = %s
                  AND c.is_active = 1
                ORDER BY u.sale_name, c.timestamp DESC
            """
            
            df_detail = pd.read_sql_query(query, self.pg_engine, params=(month_idx, year_val))
            
            if df_detail.empty:
                messagebox.showinfo("แจ้งเตือน", 
                    f"ไม่มีข้อมูลการตีกลับในรอบ {self.chart_month_var.get()} {self.chart_year_var.get()}")
                return
            
            # สรุปจำนวนครั้งต่อคน
            summary_df = df_detail.groupby('ชื่อเซลล์').size().reset_index(name='จำนวนครั้งที่ถูกตีกลับ')
            summary_df = summary_df.sort_values('จำนวนครั้งที่ถูกตีกลับ', ascending=False)
            
            # เลือกที่บันทึกไฟล์
            default_filename = f"รายงานการตีกลับ_{self.chart_month_var.get()}_{self.chart_year_var.get()}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                title="บันทึกรายงานการตีกลับ"
            )
            
            if not file_path:
                return  # ยกเลิก
            
            # สร้าง Excel Writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet 1: สรุป
                summary_df.to_excel(writer, sheet_name='สรุป', index=False)
                
                # Sheet 2: รายละเอียดทั้งหมด
                df_detail.to_excel(writer, sheet_name='รายละเอียด', index=False)
                
                # ปรับความกว้างคอลัมน์
                workbook = writer.book
                
                # Sheet สรุป
                worksheet1 = writer.sheets['สรุป']
                worksheet1.column_dimensions['A'].width = 25
                worksheet1.column_dimensions['B'].width = 25
                
                # Sheet รายละเอียด
                worksheet2 = writer.sheets['รายละเอียด']
                worksheet2.column_dimensions['A'].width = 25  # ชื่อเซลล์
                worksheet2.column_dimensions['B'].width = 20  # SO
                worksheet2.column_dimensions['C'].width = 30  # ลูกค้า
                worksheet2.column_dimensions['D'].width = 18  # ยอดขาย
                worksheet2.column_dimensions['E'].width = 50  # เหตุผล
                worksheet2.column_dimensions['F'].width = 20  # วันที่
                
                # จัดรูปแบบ Header
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                
                for ws in [worksheet1, worksheet2]:
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            messagebox.showinfo("สำเร็จ", 
                f"Export ข้อมูลเรียบร้อย!\n\nจำนวนรายการ: {len(df_detail)} รายการ\nบันทึกที่: {file_path}")
            
            # เปิดไฟล์
            if messagebox.askyesno("เปิดไฟล์", "ต้องการเปิดไฟล์ Excel หรือไม่?"):
                os.startfile(file_path)
                
        except Exception as e:
            messagebox.showerror("Error", f"Export Failed: {str(e)}")
            print(f"Export Error: {traceback.format_exc()}")

    def _load_approval_data(self):
        """ดึงรายการรออนุมัติ พร้อมระบบค้นหา"""
        for widget in self.approval_results_frame.winfo_children(): 
            widget.destroy()
        
        search_txt = self.approval_search_var.get().strip().upper()
        
        try:
            # Base Query
            query = """
                SELECT c.id, c.so_number, c.customer_name, c.sale_key, c.status, u.sale_name, c.sales_service_amount
                FROM commissions c
                LEFT JOIN sales_users u ON c.sale_key = u.sale_key
                WHERE c.status = 'Pending Sale Manager Approval' AND c.is_active = 1
            """
            params = []

            # ✅ ถ้ามีการพิมพ์ค้นหา ให้เพิ่มเงื่อนไข SQL
            if search_txt:
                term = search_txt.replace("SO", "")
                query += " AND (c.so_number ILIKE %s OR c.customer_name ILIKE %s)"
                params.extend([f"%{term}%", f"%{term}%"])

            query += " ORDER BY c.timestamp ASC"
            
            df = pd.read_sql_query(query, self.pg_engine, params=tuple(params))

            if df.empty:
                # Empty State
                empty_frame = CTkFrame(self.approval_results_frame, fg_color="transparent")
                empty_frame.pack(expand=True, pady=50)
                
                icon = "✅" if not search_txt else "🔍"
                CTkLabel(empty_frame, text=icon, font=CTkFont(size=48)).pack(pady=(0, 10))
                
                msg = "ไม่มีรายการรออนุมัติในขณะนี้" if not search_txt else "ไม่พบรายการที่ตรงกับคำค้นหา"
                CTkLabel(empty_frame, text=msg,
                        font=CTkFont(size=14),
                        text_color="#6B7280").pack()
                return

            for _, row in df.iterrows():
                self._create_so_card(self.approval_results_frame, row.to_dict(), is_approval_mode=True)
                
        except Exception as e:
            print(f"Load Approval Error: {e}")
            CTkLabel(self.approval_results_frame,
                    text=f"⚠️ เกิดข้อผิดพลาด: {str(e)[:100]}",
                    font=CTkFont(size=12),
                    text_color="#DC2626").pack(pady=20)

    # =========================================================================
    # TAB 1: DAILY REPORT WIDGET
    # =========================================================================
    def _create_daily_report_widget(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)
        self.daily_report_widget = DailyReportWidget(parent, self.app_container)
        self.daily_report_widget.pack(fill="both", expand=True)

    # =========================================================================
    # TAB 2: MASTER EDIT & SEARCH
    # =========================================================================
    def _create_master_tab(self, parent_tab):
        parent_tab.grid_columnconfigure(0, weight=1)
        parent_tab.grid_rowconfigure(1, weight=1)

        filter_frame = CTkFrame(parent_tab)
        filter_frame.grid(row=0, column=0, padx=10, pady=(10, 5), sticky="ew")
        
        current_year = datetime.now().year
        self.mst_year_var = tk.StringVar(value="ทุกปี")
        self.mst_month_var = tk.StringVar(value="ทุกเดือน")
        self.mst_day_var = tk.StringVar(value="ทุกวัน")
        self.mst_sale_var = tk.StringVar(value="All Sales")
        
        years = ["ทุกปี"] + [str(y) for y in range(current_year, current_year - 3, -1)]
        months = ["ทุกเดือน"] + ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", 
                                  "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
        days = ["ทุกวัน"] + [str(d) for d in range(1, 32)]
        sales_list = self._get_sale_list()

        CTkLabel(filter_frame, text="ปี:", font=self.entry_font).pack(side="left", padx=(10, 2))
        CTkOptionMenu(filter_frame, variable=self.mst_year_var, values=years, width=75, command=self._load_master_data).pack(side="left", padx=5)

        CTkLabel(filter_frame, text="เดือน:", font=self.entry_font).pack(side="left", padx=(10, 2))
        CTkOptionMenu(filter_frame, variable=self.mst_month_var, values=months, width=110, command=self._load_master_data).pack(side="left", padx=5)

        CTkLabel(filter_frame, text="วัน:", font=self.entry_font).pack(side="left", padx=(10, 2))
        CTkOptionMenu(filter_frame, variable=self.mst_day_var, values=days, width=70, command=self._load_master_data).pack(side="left", padx=5)

        CTkLabel(filter_frame, text="Sale:", font=self.entry_font).pack(side="left", padx=(10, 2))
        CTkOptionMenu(filter_frame, variable=self.mst_sale_var, values=sales_list, width=120, command=self._load_master_data).pack(side="left", padx=5)

        self.sm_master_search_entry = CTkEntry(filter_frame, font=self.entry_font, placeholder_text="SO / ลูกค้า...", width=150)
        self.sm_master_search_entry.pack(side="left", padx=(15, 5), fill="x", expand=True)
        self.sm_master_search_entry.bind("<Return>", lambda e: self._load_master_data())
        
        CTkButton(filter_frame, text="🔍 ค้นหา", width=80, command=self._load_master_data).pack(side="left", padx=5)
        CTkButton(filter_frame, text="↺ รีเซ็ต", width=60, fg_color="gray", command=self._reset_master_filter).pack(side="left", padx=5)

        self.sm_master_results_frame = CTkScrollableFrame(parent_tab, label_text="รายการ SO ทั้งหมด (จัดการประวัติ)")
        self.sm_master_results_frame.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        self.sm_master_results_frame.grid_columnconfigure(0, weight=1)

        self.after(200, self._load_master_data)

    def _get_sale_list(self):
        try:
            df = pd.read_sql_query("SELECT sale_key FROM sales_users WHERE role='Sale' ORDER BY sale_key", self.pg_engine)
            return ["All Sales"] + df['sale_key'].tolist()
        except: return ["All Sales"]

    def _reset_master_filter(self):
        self.mst_year_var.set("ทุกปี")
        self.mst_month_var.set("ทุกเดือน")
        self.mst_day_var.set("ทุกวัน")
        self.mst_sale_var.set("All Sales")
        self.sm_master_search_entry.delete(0, "end")
        self._load_master_data()

    def _load_master_data(self, event=None):
        for widget in self.sm_master_results_frame.winfo_children(): widget.destroy()
        try:
            query = """
                SELECT c.id, c.so_number, c.customer_name, c.sale_key, c.status, u.sale_name, c.sales_service_amount
                FROM commissions c
                LEFT JOIN sales_users u ON c.sale_key = u.sale_key
                WHERE c.is_active = 1
            """
            params = []
            if self.mst_year_var.get() != "ทุกปี":
                query += " AND EXTRACT(YEAR FROM c.timestamp::timestamp) = %s"
                params.append(int(self.mst_year_var.get()))
            if self.mst_month_var.get() != "ทุกเดือน":
                thai_months = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
                params.append(thai_months.index(self.mst_month_var.get()) + 1)
                query += " AND EXTRACT(MONTH FROM c.timestamp::timestamp) = %s"
            if self.mst_day_var.get() != "ทุกวัน":
                query += " AND EXTRACT(DAY FROM c.timestamp::timestamp) = %s"; params.append(int(self.mst_day_var.get()))
            if self.mst_sale_var.get() != "All Sales":
                query += " AND c.sale_key = %s"; params.append(self.mst_sale_var.get())
            
            search_txt = self.sm_master_search_entry.get().strip().upper()
            if search_txt:
                term = search_txt.replace("SO", "")
                query += " AND (c.so_number ILIKE %s OR c.customer_name ILIKE %s)"
                params.extend([f"%{term}%", f"%{term}%"])

            query += " ORDER BY c.timestamp DESC LIMIT 20" 
            df = pd.read_sql_query(query, self.pg_engine, params=tuple(params))

            if df.empty:
                CTkLabel(self.sm_master_results_frame, text="ไม่พบข้อมูล").pack(pady=20)
                return
            for _, row in df.iterrows():
                self._create_so_card(self.sm_master_results_frame, row.to_dict())
        except Exception as e: messagebox.showerror("Error", str(e))

    # =========================================================================
    # SHARED: SO CARD & ACTIONS
    # =========================================================================
    def _create_so_card(self, parent, so_data, is_approval_mode=False):
        """สร้าง Card แสดงข้อมูล SO พร้อมปุ่มต่างๆ"""
        so_id = so_data['id']
        so_number = so_data['so_number']
        status = so_data.get('status', 'N/A')
        amount = so_data.get('sales_service_amount', 0) or 0

        status_colors = {
            'PO In Progress': '#E0F2FE', 'Approved': '#DCFCE7', 'Paid': '#D1FAE5',
            'Rejected by SM': '#FEE2E2', 'Cancelled': '#F3F4F6', 'Draft': '#FEF3C7',
            'Pending Sale Manager Approval': '#FEF9C3'
        }
        bg_color = status_colors.get(status, "#FFFFFF")

        card = CTkFrame(parent, border_width=2, border_color="#E5E7EB",
                       corner_radius=10, fg_color=bg_color, height=70)
        card.pack(fill="x", padx=8, pady=6)
        card.pack_propagate(False)
        
        # ข้อมูล
        info_frame = CTkFrame(card, fg_color="transparent")
        info_frame.pack(side="left", fill="both", expand=True, padx=15, pady=10)
        
        line1 = f"📋 SO: {so_number}  |  👤 {so_data.get('customer_name', '-')}"
        CTkLabel(info_frame, text=line1, 
                font=CTkFont(size=13, weight="bold"),
                anchor="w").pack(anchor="w")
        
        line2 = f"🎯 เซลล์: {so_data.get('sale_name', '-')}  |  💰 {amount:,.2f} บาท  |  📌 {status}"
        CTkLabel(info_frame, text=line2,
                font=CTkFont(size=12),
                text_color="#6B7280",
                anchor="w").pack(anchor="w", pady=(3, 0))

        # ปุ่ม
        btn_frame = CTkFrame(card, fg_color="transparent")
        btn_frame.pack(side="right", padx=12, pady=10)

        if is_approval_mode:
            CTkButton(btn_frame, text="✅ อนุมัติ", width=85, height=32,
                      fg_color="#16A34A", hover_color="#15803D",
                      font=CTkFont(size=12, weight="bold"),
                      command=lambda: self._approve_so(so_id, so_number)).pack(side="left", padx=3)

        CTkButton(btn_frame, text="🛠️ แก้ไข", width=80, height=32,
                  fg_color="#4F46E5", hover_color="#4338CA",
                  font=CTkFont(size=12, weight="bold"),
                  command=lambda: self._open_so_editor_for_sm(so_number)).pack(side="left", padx=3)

        if status not in ['Cancelled', 'Rejected by SM']:
            CTkButton(btn_frame, text="❌ ตีกลับ", width=80, height=32,
                      fg_color="#DC2626", hover_color="#B91C1C",
                      font=CTkFont(size=12, weight="bold"),
                      command=lambda: self._reject_so(so_id, so_number)).pack(side="left", padx=3)

    # ✅ ฟังก์ชันอนุมัติ SO
    def _approve_so(self, so_id, so_number):
        if not messagebox.askyesno("ยืนยัน", f"คุณต้องการอนุมัติ SO: {so_number} ใช่หรือไม่?"):
            return

        conn = None
        try:
            conn = self.app_container.get_connection()
            with conn.cursor() as cursor:
                cursor.execute("""
                    UPDATE commissions 
                    SET status = 'Pending PU', 
                        approver_sale_manager_key = %s, 
                        approval_date_sale_manager = CURRENT_TIMESTAMP,
                        claim_timestamp = NULL
                    WHERE id = %s
                """, (self.user_key, so_id))
                
                cursor.execute("SELECT sale_key FROM sales_users WHERE role = 'Purchasing Staff' AND status = 'Active'")
                pu_keys = [row[0] for row in cursor.fetchall()]
                
                for pu_key in pu_keys:
                    cursor.execute("""
                        INSERT INTO notifications (user_key_to_notify, message, is_read, related_so_id)
                        VALUES (%s, %s, FALSE, %s)
                    """, (pu_key, f"มี SO ใหม่ ({so_number}) ผ่านการอนุมัติแล้ว รอคุณ Claim งาน", so_id))

            conn.commit()
            messagebox.showinfo("สำเร็จ", f"อนุมัติ SO: {so_number} เรียบร้อยแล้ว")
            self._refresh_all_tabs()
            
        except Exception as e:
            if conn: conn.rollback()
            messagebox.showerror("Error", f"Approve Failed: {e}")
        finally:
            if conn: self.app_container.release_connection(conn)

    def _reject_so(self, so_id, so_number):
        """เปิดหน้าต่างเลือกเหตุผลการตีกลับ"""
        
        def save_rejection(reason):
            conn = None
            try:
                conn = self.app_container.get_connection()
                with conn.cursor() as cursor:
                    cursor.execute("""
                        UPDATE commissions 
                        SET status = 'Rejected by SM', 
                            rejection_reason = %s 
                        WHERE id = %s
                    """, (reason, so_id))
                    
                    cursor.execute("SELECT sale_key FROM commissions WHERE id = %s", (so_id,))
                    res = cursor.fetchone()
                    
                    if res:
                        cursor.execute("""
                            INSERT INTO notifications (user_key_to_notify, message, is_read, related_so_id) 
                            VALUES (%s, %s, FALSE, %s)
                        """, (res[0], f"SO: {so_number} ถูกตีกลับ: {reason}", so_id))
                
                conn.commit()
                messagebox.showinfo("สำเร็จ", f"ตีกลับ SO: {so_number} เรียบร้อยแล้ว")
                self._refresh_all_tabs()
                
            except Exception as e:
                if conn: conn.rollback()
                messagebox.showerror("Error", f"Reject Failed: {e}")
            finally:
                if conn: self.app_container.release_connection(conn)

        SORejectionDialog(self, so_number, save_rejection)

    def _open_so_editor_for_sm(self, so_number):
        if self.so_popup is not None and self.so_popup.winfo_exists():
            self.so_popup.focus()
            return
        try:
            so_df = pd.read_sql_query("SELECT * FROM commissions WHERE so_number = %s AND is_active = 1 LIMIT 1", 
                                     self.pg_engine, params=(so_number,))
            if so_df.empty:
                messagebox.showerror("Error", "ไม่พบข้อมูล SO")
                return
            
            def _refresh_on_save():
                self._refresh_all_tabs()
            
            self.so_popup = SOPopupWindow(
                master=self,
                app_container=self.app_container,
                sales_data=so_df.iloc[0].to_dict(),
                so_shared_vars=self.so_shared_vars,
                sale_theme=self.sale_theme,
                on_save_callback=_refresh_on_save
            )
        except Exception as e:
            messagebox.showerror("Error", f"Open Editor Failed: {e}")
            print(traceback.format_exc())

    def _so_create_string_vars(self):
        """สร้าง StringVars สำหรับหน้าจอแก้ไข SO"""
        self.so_shared_vars = {}
        now = datetime.now()
        thai_months_list = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", 
                            "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
        
        self.so_shared_vars.update({
            'thai_months': thai_months_list,
            'thai_month_map': {name: i + 1 for i, name in enumerate(thai_months_list)},
            'customer_type_var': tk.StringVar(value="ลูกค้าเก่า"),
            'credit_term_var': tk.StringVar(value="เงินสด"),
            'commission_month_var': tk.StringVar(value=thai_months_list[now.month - 1]),
            'commission_year_var': tk.StringVar(value=str(now.year + 543)),
            'payment1_percent_var': tk.StringVar(value="ระบุยอดเอง"),
            'payment2_percent_var': tk.StringVar(value="ระบุยอดเอง"),
            'delivery_type_var': tk.StringVar(value="ซัพพลายเออร์จัดส่ง"),
            'payment_total_var': tk.StringVar(value="0.00"),
            'so_subtotal_var': tk.StringVar(value="0.00"),
            'so_vat_var': tk.StringVar(value="0.00"),
            'so_grand_total_var': tk.StringVar(value="0.00"),
            'so_vs_payment_result_var': tk.StringVar(value="-"),
            'difference_amount_var': tk.StringVar(value="0.00"),
            'balance_due_var': tk.StringVar(value="0.00"),
            'cash_product_input_var': tk.StringVar(value="0.00"),
            'cash_service_total_var': tk.StringVar(value="0.00"),
            'cash_required_total_var': tk.StringVar(value="0.00"),
            'cash_actual_payment_var': tk.StringVar(value="0.00"),
            'cash_verification_result_var': tk.StringVar(value="-"),
            'sales_vat_calc_var': tk.StringVar(value="0.00"),
            'cutting_drilling_vat_calc_var': tk.StringVar(value="0.00"),
            'other_service_vat_calc_var': tk.StringVar(value="0.00"),
            'shipping_vat_calc_var': tk.StringVar(value="0.00"),
            'card_fee_vat_calc_var': tk.StringVar(value="0.00"),
            'relocation_vat_calc_var': tk.StringVar(value="0.00"),
            'sales_service_vat_option': tk.StringVar(value="VAT"),
            'cutting_drilling_fee_vat_option': tk.StringVar(value="VAT"),
            'other_service_fee_vat_option': tk.StringVar(value="VAT"),
            'shipping_vat_option_var': tk.StringVar(value="VAT"),
            'credit_card_fee_vat_option_var': tk.StringVar(value="VAT"),
            'relocation_cost_vat_option': tk.StringVar(value="VAT")
        })


class SORejectionDialog(CTkToplevel):
    def __init__(self, master, so_number, on_confirm_callback):
        super().__init__(master)
        self.title(f"ตีกลับ SO: {so_number}")
        self.geometry("450x550")
        self.on_confirm_callback = on_confirm_callback
        
        self.grid_columnconfigure(0, weight=1)
        self.attributes("-topmost", True) 

        CTkLabel(self, text=f"ระบุเหตุผลที่ตีกลับ SO: {so_number}", 
                font=CTkFont(size=16, weight="bold")).pack(pady=15)

        self.reasons = [
            "เลขที่ใบสั่งขาย (SO) ไม่ถูกต้อง",
            "ข้อมูลชื่อลูกค้าไม่ถูกต้อง",
            "ค่าจัดส่ง / ค่าขนส่งไม่ถูกต้อง",
            "ยอดโอนชำระไม่ถูกต้อง (ไม่ตรงตามสลิป)",
            "ยอดขายสินค้าหรือค่าบริการไม่ถูกต้อง",
            "วันที่จัดส่งสินค้าไม่ถูกต้อง"
        ]
        
        self.check_vars = []
        container = CTkFrame(self, fg_color="transparent")
        container.pack(fill="x", padx=30)

        for reason in self.reasons:
            var = tk.BooleanVar(value=False)
            cb = CTkCheckBox(container, text=reason, variable=var, font=CTkFont(size=13))
            cb.pack(anchor="w", pady=5)
            self.check_vars.append((var, reason))

        CTkLabel(self, text="อื่นๆ / ระบุเพิ่มเติม:", 
                font=CTkFont(size=13, weight="bold")).pack(anchor="w", padx=30, pady=(15, 5))
        self.other_reason_entry = CTkEntry(self, placeholder_text="พิมพ์เหตุผลเพิ่มเติมที่นี่...", width=380)
        self.other_reason_entry.pack(padx=30, pady=(0, 20))

        btn_frame = CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=30, pady=10)
        
        CTkButton(btn_frame, text="ยกเลิก", fg_color="gray", width=100, 
                 command=self.destroy).pack(side="left", padx=5)
        CTkButton(btn_frame, text="ตกลง (ตีกลับ)", fg_color="#DC2626", hover_color="#B91C1C", 
                 width=100, command=self._on_confirm).pack(side="right", padx=5)

    def _on_confirm(self):
        selected_reasons = [text for var, text in self.check_vars if var.get()]
        other_text = self.other_reason_entry.get().strip()
        
        if other_text:
            selected_reasons.append(other_text)

        if not selected_reasons:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกหรือระบุเหตุผลอย่างน้อย 1 ข้อ", parent=self)
            return

        final_reason = ", ".join(selected_reasons)
        self.on_confirm_callback(final_reason)
        self.destroy()