import tkinter as tk
from tkinter import messagebox
from customtkinter import (
    CTkFrame, CTkLabel, CTkFont, CTkButton,
    CTkOptionMenu, CTkScrollableFrame, CTkComboBox  # 🟢 เพิ่ม CTkComboBox เข้ามา
)
import pandas as pd
import psycopg2.extras
from datetime import datetime
from tkcalendar import DateEntry, Calendar

class PatchedDateEntry(DateEntry):
    """แก้ปัญหา calendar ปิดตัวเองเมื่อกดปุ่มเลื่อนเดือน/ปี"""
    def _on_focus_out_cal(self, event):
        # ตรวจสอบว่า focus ไปอยู่ที่ widget ลูก (ปุ่มใน calendar) หรือไม่
        if event.widget in self._top_cal.winfo_children():
            return
        # ตรวจสอบลึกลงไปอีกชั้น (ปุ่มลูกศรอาจเป็น grandchild)
        try:
            focused = self._top_cal.focus_get()
            if focused and str(focused).startswith(str(self._top_cal)):
                return
        except Exception:
            pass
        super()._on_focus_out_cal(event)

    def _drop_down(self):
        super()._drop_down()
        # rebind FocusOut ของ Calendar ภายในด้วย
        if hasattr(self, '_top_cal') and self._top_cal:
            self._top_cal.bind("<FocusOut>", self._on_focus_out_cal)

try:
    from tksheet import Sheet
    HAS_TKSHEET = True
except ImportError:
    HAS_TKSHEET = False


# ============================================================
# COLOR PALETTE (Power BI style)
# ============================================================
COLORS = {
    "bg_main":         "#F3F4F6",
    "bg_sidebar":      "#FFFFFF",
    "bg_white":        "#FFFFFF",
    "header_blue":     "#1D4ED8",   # dark-blue table header
    "header_text":     "#FFFFFF",
    "row_alt":         "#F8FAFC",
    "row_normal":      "#FFFFFF",
    "highlight_green": "#D1FAE5",
    "highlight_green_fg": "#065F46",
    "highlight_red":   "#FEE2E2",
    "highlight_red_fg": "#991B1B",
    "kpi_blue":        "#1D4ED8",
    "kpi_green":       "#047857",
    "kpi_red":         "#B91C1C",
    "kpi_purple":      "#6D28D9",
    "border":          "#E5E7EB",
    "text_dark":       "#111827",
    "text_medium":     "#4B5563",
    "text_light":      "#9CA3AF",
    "filter_bg":       "#F9FAFB",
    "filter_border":   "#D1D5DB",
    "btn_blue":        "#2563EB",
    "btn_hover":       "#1D4ED8",
    "sidebar_header":  "#1E3A5F",
}

FONT_FAMILY = "Tahoma"


class DashboardCostScreen(CTkFrame):
    def __init__(self, master, app_container):
        super().__init__(master, fg_color=COLORS["bg_main"])
        self.app_container = app_container
        self.current_user = getattr(self.app_container, 'current_user_key', 'PU_Default')

        self.grid_columnconfigure(0, weight=0, minsize=220)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.raw_df = pd.DataFrame()
        self.current_filtered_df = pd.DataFrame()   # เก็บ filtered ล่าสุดสำหรับ Export
        self.all_order_nos = []
        self.all_sale_order_nos = []
        self.all_product_names = []
        self.all_suppliers = []

        self._build_sidebar()
        self._build_main_content()

        self.after(200, self._load_data_from_db)

    # =========================================================
    # SIDEBAR
    # =========================================================
    def _build_sidebar(self):
        sidebar = CTkScrollableFrame(self, fg_color=COLORS["bg_sidebar"], corner_radius=0,
                                    border_width=0)
        sidebar.grid(row=0, column=0, sticky="nsew")

        header_band = CTkFrame(sidebar, fg_color=COLORS["sidebar_header"], corner_radius=0, height=50)
        header_band.pack(fill="x", pady=(0, 10))
        header_band.pack_propagate(False)
        CTkLabel(header_band, text="  ⚙  Filters",
                font=CTkFont(family=FONT_FAMILY, size=14, weight="bold"),
                text_color="#FFFFFF").pack(side="left", padx=10, pady=12)

        self.filter_vars = {
            "date_from":     tk.StringVar(value=""),
            "date_to":       tk.StringVar(value=""),
            "order_no":      tk.StringVar(value="All"),
            "sale_order_no": tk.StringVar(value="All"),
            "product_name":  tk.StringVar(value="All"),
            "pu_user":       tk.StringVar(value="All"),
            "sale_name":     tk.StringVar(value="All"),
            "supplier":      tk.StringVar(value="All"),
            "status":        tk.StringVar(value="All"),
            "priority":      tk.StringVar(value="All"),
            "select":        tk.StringVar(value="All"),
            "year":          tk.StringVar(value="All"),
            "month":         tk.StringVar(value="All"),
        }
        self.filter_menus = {}

        date_frame = CTkFrame(sidebar, fg_color="transparent")
        date_frame.pack(fill="x", padx=12, pady=(10, 5))
        
        CTkLabel(date_frame, text="📅 กรองตามช่วงวันที่ขอราคา", 
                font=CTkFont(family=FONT_FAMILY, size=12, weight="bold"), 
                text_color=COLORS["text_medium"]).pack(anchor="w", pady=(0, 8))
                
        # ✅ เปลี่ยนเป็น PatchedDateEntry (แก้ปัญหา calendar ปิดตัวเองเมื่อกดลูกศรเลื่อนเดือน/ปี)
        from_frame = CTkFrame(date_frame, fg_color="transparent")
        from_frame.pack(fill="x", pady=(0, 5))
        CTkLabel(from_frame, text="ตั้งแต่:", font=CTkFont(family=FONT_FAMILY, size=11), width=40, anchor="w").pack(side="left")
        
        self.cal_from = PatchedDateEntry(          # ✅ เปลี่ยนจาก DateEntry → PatchedDateEntry
            from_frame, width=15,
            background=COLORS["header_blue"],
            foreground='white', borderwidth=0,
            headersbackground=COLORS["header_blue"], headersforeground='white',
            selectbackground=COLORS["btn_blue"], selectforeground='white',
            normalbackground=COLORS["bg_white"], normalforeground=COLORS["text_dark"],
            weekendbackground=COLORS["bg_white"], weekendforeground=COLORS["highlight_red_fg"],
            othermonthforeground=COLORS["text_light"], othermonthbackground=COLORS["bg_white"],
            othermonthweforeground=COLORS["text_light"], othermonthwebackground=COLORS["bg_white"],
            font=(FONT_FAMILY, 14),
            date_pattern='dd/mm/yyyy', 
            locale='th_TH',
            showweeknumbers=False,
            showothermonthdays=False,
            year=datetime.now().year,
            textvariable=self.filter_vars["date_from"]
        )
        self.cal_from.pack(side="left", fill="x", expand=True)
        self.cal_from.delete(0, 'end')
        
        to_frame = CTkFrame(date_frame, fg_color="transparent")
        to_frame.pack(fill="x", pady=(0, 5))
        CTkLabel(to_frame, text="ถึง:", font=CTkFont(family=FONT_FAMILY, size=11), width=40, anchor="w").pack(side="left")
        
        self.cal_to = PatchedDateEntry(            # ✅ เปลี่ยนจาก DateEntry → PatchedDateEntry
            to_frame, width=15,
            background=COLORS["header_blue"],
            foreground='white', borderwidth=0,
            headersbackground=COLORS["header_blue"], headersforeground='white',
            selectbackground=COLORS["btn_blue"], selectforeground='white',
            normalbackground=COLORS["bg_white"], normalforeground=COLORS["text_dark"],
            weekendbackground=COLORS["bg_white"], weekendforeground=COLORS["highlight_red_fg"],
            othermonthforeground=COLORS["text_light"], othermonthbackground=COLORS["bg_white"],
            othermonthweforeground=COLORS["text_light"], othermonthwebackground=COLORS["bg_white"],
            font=(FONT_FAMILY, 14),
            date_pattern='dd/mm/yyyy',
            locale='th_TH',
            showweeknumbers=False,
            showothermonthdays=False,
            year=datetime.now().year,
            textvariable=self.filter_vars["date_to"]
        )
        self.cal_to.pack(side="left", fill="x", expand=True)
        self.cal_to.delete(0, 'end')

        self.cal_from.bind("<<DateEntrySelected>>", self._apply_filters)
        self.cal_to.bind("<<DateEntrySelected>>", self._apply_filters)

        filters_config = [
            ("🔍 Order No. (พิมพ์ค้นหา)", "order_no"),
            ("🔍 Sale Order No. (พิมพ์ค้นหา)", "sale_order_no"),
            ("🔍 รายการสินค้า (พิมพ์ค้นหา)", "product_name"),
            ("🔍 ชื่อ Supplier (พิมพ์ค้นหา)", "supplier"),  # ← เหลือแค่นี้อันเดียว
            ("ผู้ทำตาราง (PU User)", "pu_user"),
            ("ชื่อ Sale",        "sale_name"),
            # ← ลบ "ชื่อ Supplier" ตัวที่สองออก
            ("สถานะ",            "status"),
            ("PRIORITY",        "priority"),
            ("Select",          "select"),
            ("ปี (Year)",        "year"),
            ("เดือน (Month)",     "month"),
        ]

        for label_text, key in filters_config:
            lbl = CTkLabel(sidebar, text=label_text,
                        font=CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
                        text_color=COLORS["text_medium"])
            lbl.pack(anchor="w", padx=12, pady=(8, 1))

            if key in ["order_no", "sale_order_no", "product_name", "supplier"]:
                menu = CTkComboBox(
                    sidebar,
                    variable=self.filter_vars[key],
                    values=["All"],
                    width=190, height=30,
                    fg_color=COLORS["bg_white"],
                    text_color=COLORS["text_dark"],
                    border_color=COLORS["filter_border"],
                    button_color=COLORS["filter_border"],
                    button_hover_color=COLORS["header_blue"],
                    dropdown_fg_color=COLORS["bg_white"],
                    dropdown_text_color=COLORS["text_dark"],
                    font=CTkFont(family=FONT_FAMILY, size=12),
                    command=self._apply_filters
                )
                menu.pack(fill="x", padx=12, pady=(0, 2))
                
                if key == "order_no":
                    menu.bind("<KeyRelease>", self._on_order_search)
                elif key == "sale_order_no":
                    menu.bind("<KeyRelease>", self._on_sale_order_search)
                elif key == "product_name":
                    menu.bind("<KeyRelease>", self._on_product_search)
                elif key == "supplier": 
                    menu.bind("<KeyRelease>", self._on_supplier_search) # 🟢 เพิ่มบรรทัดนี้
                    
                menu.bind("<Return>", self._apply_filters)
                
            else:
                menu = CTkOptionMenu(
                    sidebar,
                    variable=self.filter_vars[key],
                    values=["All"],
                    width=190, height=30,
                    fg_color=COLORS["filter_bg"],
                    text_color=COLORS["text_dark"],
                    button_color=COLORS["filter_border"],
                    button_hover_color=COLORS["header_blue"],
                    dropdown_fg_color=COLORS["bg_white"],
                    dropdown_text_color=COLORS["text_dark"],
                    font=CTkFont(family=FONT_FAMILY, size=12),
                    command=self._apply_filters,
                )
                menu.pack(fill="x", padx=12, pady=(0, 2))
            
            self.filter_menus[key] = menu

        CTkFrame(sidebar, fg_color=COLORS["border"], height=1).pack(fill="x", padx=12, pady=15)

        CTkButton(
            sidebar, text="🔄  รีเฟรชฐานข้อมูล",
            fg_color=COLORS["btn_blue"], hover_color=COLORS["btn_hover"],
            font=CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            corner_radius=6, height=36,
            command=self._load_data_from_db
        ).pack(fill="x", padx=12, pady=(0, 8))

        CTkButton(
            sidebar, text="📥  Export Excel",
            fg_color="#16A34A", hover_color="#15803D",
            font=CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            corner_radius=6, height=36,
            command=self._export_to_excel
        ).pack(fill="x", padx=12, pady=(0, 20))

    # =========================================================
    # EXPORT TO EXCEL
    # =========================================================
    def _export_to_excel(self):
        from tkinter import filedialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        df = self.current_filtered_df
        if df.empty:
            messagebox.showwarning("ไม่มีข้อมูล",
                "ยังไม่มีข้อมูลที่จะ Export\nกรุณาโหลดหรือ Filter ข้อมูลก่อน",
                parent=self)
            return

        # ── เลือกที่บันทึก ────────────────────────────────────────
        from datetime import datetime
        default_name = f"CostBenchmark_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_name,
            title="บันทึกไฟล์ Excel",
            parent=self
        )
        if not save_path:
            return

        try:
            # ── ซ่อนคอลัมน์ internal ที่ไม่ต้องการ export ────────────
            hide_cols = {"benchmark_year", "benchmark_month"}
            export_df = df[[c for c in df.columns if c not in hide_cols]].copy()

            # ── helper: แทรกคอลัมน์หลัง anchor ──────────────────────────
            def _insert_after(df, anchor_col, new_cols):
                cols = list(df.columns)
                for cc in new_cols:          # ลบออกจากตำแหน่งเดิม (ถ้ามี)
                    if cc in cols:
                        cols.remove(cc)
                idx = cols.index(anchor_col) + 1 if anchor_col in cols else len(cols)
                cols = cols[:idx] + new_cols + cols[idx:]
                return df[cols]

            # 🟢 คอลัมน์ 1: ส่วนลดรวม 1+2 และ % — แทรกหลัง "ส่วนลด 2 (บาท)"
            try:
                d1  = pd.to_numeric(export_df.get("ส่วนลด 1 (บาท)", 0), errors='coerce').fillna(0)
                d2  = pd.to_numeric(export_df.get("ส่วนลด 2 (บาท)", 0), errors='coerce').fillna(0)
                qty = pd.to_numeric(export_df.get("จำนวน", 0), errors='coerce').fillna(0)
                disc_sum = (d1 + d2) * qty
                export_df["ส่วนลดรวม 1+2"]     = disc_sum.replace(0, float('nan'))
                tunkruam = pd.to_numeric(export_df.get("ทุนรวม", 0), errors='coerce').fillna(0)
                pct_disc = (disc_sum / tunkruam.replace(0, float('nan')) * 100).round(2)
                export_df["ส่วนลดรวม 1+2 (%)"] = pct_disc.replace(0, float('nan'))
                export_df = _insert_after(export_df, "ส่วนลด 2 (บาท)",
                                          ["ส่วนลดรวม 1+2", "ส่วนลดรวม 1+2 (%)"])
            except Exception as e:
                print(f"[Export] คำนวณ ส่วนลดรวม error: {e}")

            # 🟢 คอลัมน์ 2: avg. cost/per order และ Price competitiveness
            # — แทรกหลัง "ต้นทุนรวม (รวมย้าย)"
            try:
                if ("รายการสินค้า" in export_df.columns
                        and "ต้นทุนรวม (รวมย้าย)" in export_df.columns
                        and "Select" in export_df.columns):
                    _EXP_SEL_TYPES  = {"✔", "เทียบเพื่อชุบ ✔"}
                    _EXP_CHUB_TYPES = {"เทียบเพื่อชุบ", "เทียบเพื่อชุบ ✔"}

                    _ev = export_df[export_df["รายการสินค้า"].astype(str).str.strip() != ""].copy()
                    _ev["ต้นทุนรวม (รวมย้าย)"] = pd.to_numeric(
                        _ev["ต้นทุนรวม (รวมย้าย)"], errors='coerce')
                    _ev = _ev[_ev["ต้นทุนรวม (รวมย้าย)"] > 0].copy()
                    _ev["_grp"] = _ev["Select"].astype(str).str.strip().apply(
                        lambda s: "chub" if s in _EXP_CHUB_TYPES else "normal")

                    # avg แยกตาม (SKU, กลุ่ม) + นับจำนวน supplier
                    _exp_avg_map = {}
                    _grp_avg   = _ev.groupby(["รายการสินค้า", "_grp"])["ต้นทุนรวม (รวมย้าย)"].mean()
                    _grp_count = _ev.groupby(["รายการสินค้า", "_grp"])["ต้นทุนรวม (รวมย้าย)"].count()
                    for (sku, grp), avg in _grp_avg.items():
                        _exp_avg_map[(sku, grp)] = (avg, int(_grp_count.get((sku, grp), 1)))

                    # row ที่ถูก ✔ ต่อ (SKU, grp) — ใช้แสดง avg. cost/per order
                    _exp_sel_idx = {}
                    for _ei, _er in export_df.iterrows():
                        _esel = str(_er.get("Select", "")).strip()
                        if _esel in _EXP_SEL_TYPES:
                            _esku = str(_er.get("รายการสินค้า", "")).strip()
                            _egrp = "chub" if _esel in _EXP_CHUB_TYPES else "normal"
                            _ekey = (_esku, _egrp)
                            if _esku and _ekey not in _exp_sel_idx:
                                _exp_sel_idx[_ekey] = _ei

                    avg_col   = []
                    comp_col  = []
                    _seen_avg = set()
                    for _ei, erow in export_df.iterrows():
                        sku      = str(erow.get("รายการสินค้า", "")).strip()
                        sel_type = str(erow.get("Select", "")).strip()
                        cost     = pd.to_numeric(erow.get("ต้นทุนรวม (รวมย้าย)", 0), errors='coerce')
                        grp      = "chub" if sel_type in _EXP_CHUB_TYPES else "normal"
                        map_key  = (sku, grp)
                        entry    = _exp_avg_map.get(map_key)

                        # avg. cost/per order — แสดงที่ row ✔ ของกลุ่ม
                        if entry and sku and map_key not in _seen_avg:
                            avg_v, cnt = entry
                            sel_i = _exp_sel_idx.get(map_key)
                            if sel_i == _ei or sel_i is None:
                                _seen_avg.add(map_key)
                                avg_col.append(round(avg_v, 2))
                            else:
                                avg_col.append(float('nan'))
                        else:
                            avg_col.append(float('nan'))

                        # Price competitiveness — เฉพาะ ✔ / เทียบเพื่อชุบ ✔ ที่มี 2+ supplier
                        if sel_type in _EXP_SEL_TYPES and sku and entry:
                            avg_v, cnt = entry
                            if avg_v and avg_v != 0 and cnt > 1 and cost and not pd.isna(cost):
                                comp_col.append(round((cost - avg_v) / avg_v * 100, 2))
                            else:
                                comp_col.append(float('nan'))
                        else:
                            comp_col.append(float('nan'))

                    export_df["avg. cost/per order"]    = avg_col
                    export_df["Price competitiveness"] = comp_col
                    export_df = _insert_after(export_df, "ต้นทุนรวม (รวมย้าย)",
                                              ["avg. cost/per order", "Price competitiveness"])
            except Exception as e:
                print(f"[Export] คำนวณ avg/competitiveness error: {e}")

            # ── สร้าง Workbook ────────────────────────────────────────
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cost Benchmark"

            # สี theme
            HDR_BG   = "1E3A5F"   # navy
            HDR_FG   = "FFFFFF"
            ROW_ALT  = "EFF6FF"   # ฟ้าอ่อน
            ROW_NORM = "FFFFFF"
            BORDER_C = "CBD5E1"

            thin = Side(style="thin", color=BORDER_C)
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # ── Header row ────────────────────────────────────────────
            headers = list(export_df.columns)
            for col_idx, col_name in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font      = Font(bold=True, color=HDR_FG, size=11)
                cell.fill      = PatternFill("solid", fgColor=HDR_BG)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center", wrap_text=True)
                cell.border    = border
            ws.row_dimensions[1].height = 28

            # ── กำหนด number format ต่อ column ───────────────────────────
            _num_fmt = {}   # col_idx (1-based) → format string
            for _ci, _cn in enumerate(headers, start=1):
                if _cn == "avg. cost/per order":
                    _num_fmt[_ci] = '#,##0.00'
                elif _cn == "Price competitiveness":
                    _num_fmt[_ci] = '#,##0.00"%"'   # แสดงเป็น -5.62%

            # ── Data rows ─────────────────────────────────────────────
            for row_idx, (_, row_data) in enumerate(export_df.iterrows(), start=2):
                bg = ROW_ALT if row_idx % 2 == 0 else ROW_NORM
                fill = PatternFill("solid", fgColor=bg)
                for col_idx, value in enumerate(row_data, start=1):
                    # แปลงค่า NaN → ว่าง
                    if pd.isna(value) if not isinstance(value, str) else False:
                        value = ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = Alignment(vertical="center")
                    if col_idx in _num_fmt and value != "":
                        cell.number_format = _num_fmt[col_idx]

            # ── Auto-fit column width ─────────────────────────────────
            for col_idx, col_name in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(col_name)),
                    *[len(str(r[col_idx - 1]))
                      for r in export_df.itertuples(index=False)
                      if r[col_idx - 1] is not None],
                    0
                )
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

            # ── Freeze header row ─────────────────────────────────────
            ws.freeze_panes = "A2"

            # ── Info sheet — บันทึก filter ที่ใช้ ─────────────────────
            ws_info = wb.create_sheet("Filter Info")
            ws_info.append(["Filter", "ค่าที่เลือก"])
            ws_info.append(["Export เมื่อ", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
            ws_info.append(["จำนวนแถวที่ export", len(export_df)])

            filter_label_map = {
                "date_from":      "วันที่ (จาก)",
                "date_to":        "วันที่ (ถึง)",
                "order_no":       "Order No.",
                "sale_order_no":  "Sale Order No.",
                "product_name":   "รายการสินค้า",
                "supplier":       "ชื่อ Supplier",
                "sale_name":      "ชื่อ Sale",
                "status":         "สถานะ",
                "priority":       "PRIORITY",
                "select":         "Select",
                "year":           "ปี (Year)",
                "month":          "เดือน (Month)",
                "pu_user":        "ผู้ทำตาราง (PU User)",
            }
            for key, label in filter_label_map.items():
                val = self.filter_vars.get(key, tk.StringVar()).get()
                if val and val != "All":
                    ws_info.append([label, val])

            for cell in ws_info["A"]:
                cell.font = Font(bold=True)

            wb.save(save_path)

            messagebox.showinfo(
                "Export สำเร็จ",
                f"✅ Export ข้อมูล {len(export_df):,} แถว เรียบร้อยแล้ว\n\n{save_path}",
                parent=self
            )

        except Exception as e:
            messagebox.showerror("Export ผิดพลาด", f"ไม่สามารถ Export ได้:\n{e}", parent=self)
            import traceback; traceback.print_exc()

    # 🟢 ฟังก์ชันอัจฉริยะ สำหรับค้นหา Order No. ใน Dropdown แบบเรียลไทม์
    def _on_order_search(self, event):
        if event.keysym in ['Up', 'Down', 'Left', 'Right', 'Return', 'Escape']:
            return
        typed = self.filter_vars["order_no"].get().lower()
        matches = [o for o in self.all_order_nos if typed in str(o).lower()] if typed and typed != "all" else self.all_order_nos
        self._show_search_popup("order_no", matches, self.filter_menus["order_no"])

    def _on_sale_order_search(self, event):
        if event.keysym in ['Up', 'Down', 'Left', 'Right', 'Return', 'Escape']:
            return
        typed = self.filter_vars["sale_order_no"].get().lower()
        matches = [o for o in self.all_sale_order_nos if typed in str(o).lower()] if typed and typed != "all" else self.all_sale_order_nos
        self._show_search_popup("sale_order_no", matches, self.filter_menus["sale_order_no"])

    def _on_product_search(self, event):
        if event.keysym in ['Up', 'Down', 'Left', 'Right', 'Return', 'Escape']:
            return
        typed = self.filter_vars["product_name"].get().lower()
        matches = [p for p in self.all_product_names if typed in str(p).lower()] if typed and typed != "all" else self.all_product_names
        self._show_search_popup("product_name", matches, self.filter_menus["product_name"])

    def _on_supplier_search(self, event):
        if event.keysym in ['Up', 'Down', 'Left', 'Right', 'Return', 'Escape']:
            return
        typed = self.filter_vars["supplier"].get().lower()
        matches = [s for s in self.all_suppliers if typed in str(s).lower()] if typed and typed != "all" else self.all_suppliers
        self._show_search_popup("supplier", matches, self.filter_menus["supplier"])

        

    def _show_search_popup(self, key, matches, anchor_widget):
        """แสดง popup listbox ใต้ช่อง input โดยไม่บัง"""
        # ปิด popup เก่าถ้ามี
        if hasattr(self, '_search_popup') and self._search_popup:
            try: self._search_popup.destroy()
            except: pass
            self._search_popup = None

        if not matches:
            return

        popup = tk.Toplevel(self)
        popup.overrideredirect(True)
        popup.attributes('-topmost', True)
        self._search_popup = popup

        # วางตำแหน่งใต้ anchor_widget
        anchor_widget.update_idletasks()
        x = anchor_widget.winfo_rootx()
        y = anchor_widget.winfo_rooty() + anchor_widget.winfo_height() + 2
        w = anchor_widget.winfo_width()

        popup.configure(bg="#D1D5DB")

        frame = tk.Frame(popup, bg="white")
        frame.pack(fill="both", expand=True, padx=1, pady=1)

        sb = tk.Scrollbar(frame)
        sb.pack(side="right", fill="y")

        h = min(len(matches), 8)
        lb = tk.Listbox(frame, font=(FONT_FAMILY, 11),
                        selectbackground="#3B82F6", selectforeground="white",
                        relief="flat", borderwidth=0, highlightthickness=0,
                        height=h, yscrollcommand=sb.set, activestyle="none")
        lb.pack(side="left", fill="both", expand=True)
        sb.config(command=lb.yview)

        for item in matches[:100]:
            lb.insert(tk.END, item)

        popup.geometry(f"{w}x{h * 22 + 4}+{x}+{y}")

        def on_select(e=None):
            sel = lb.curselection()
            if sel:
                val = lb.get(sel[0])
                self.filter_vars[key].set(val)
                self._apply_filters()
            try: popup.destroy()
            except: pass
            self._search_popup = None

        def on_close(e=None):
            self.after(200, lambda: _try_close())

        def _try_close():
            try:
                focused = str(self.focus_get())
                # ถ้า focus ยังอยู่ใน popup หรือ anchor → ไม่ปิด
                if str(popup) in focused or str(anchor_widget) in focused:
                    return
                popup.destroy()
                self._search_popup = None
            except:
                pass

        lb.bind("<ButtonRelease-1>", on_select)
        lb.bind("<Return>", on_select)
        lb.bind("<Escape>", lambda e: [popup.destroy().__setattr__('_search_popup', None)])
        
        # ปิดเมื่อ focus ออกจาก anchor
        anchor_widget.bind("<FocusOut>", on_close, add="+")
        anchor_widget.bind("<Escape>", lambda e: popup.destroy(), add="+")
        
        # scroll keyboard ใน listbox
        anchor_widget.bind("<Down>", lambda e: [lb.focus_set(), lb.selection_set(0)] if lb.size() > 0 else None, add="+")


    # =========================================================
    # MAIN CONTENT
    # =========================================================
    def _build_main_content(self):
        main_frame = CTkFrame(self, fg_color="transparent")
        main_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 10), pady=10)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(2, weight=1)

        # ── Header bar ──────────────────────────────────────
        header_frame = CTkFrame(main_frame, fg_color=COLORS["sidebar_header"],
                                corner_radius=8, height=48)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        header_frame.grid_propagate(False)
        CTkLabel(header_frame,
                 text="  📊  Cost Benchmark Dashboard",
                 font=CTkFont(family=FONT_FAMILY, size=18, weight="bold"),
                 text_color="#FFFFFF").pack(side="left", padx=14, pady=10)

        # ── KPI Cards ────────────────────────────────────────
        kpi_frame = CTkFrame(main_frame, fg_color="transparent")
        kpi_frame.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        
        # 🟢 [แก้ไข] เปลี่ยนจาก range(5) เป็น range(6) เพราะเรามี 6 การ์ดแล้ว
        for i in range(6):
            kpi_frame.grid_columnconfigure(i, weight=1)

        self.kpi_labels = {}
        kpi_config = [
            ("จำนวน Order รวม",        "total_orders", COLORS["kpi_blue"],   "🗂"),
            ("ยอดขายรวม",               "total_sales",  COLORS["kpi_green"],  "📈"),
            ("ยอด Win รวม",             "total_win",    "#10B981",            "🏆"),
            ("ปริมาณรวม (เส้น/ชิ้น)",  "total_qty",    COLORS["kpi_green"],  "📦"),
            ("ยอดซื้อรวม (ทุน)",        "total_cost",   COLORS["kpi_red"],    "💰"),
            ("Markup เฉลี่ย",           "avg_margin",   COLORS["kpi_purple"], "📊"),
        ]

        for i, (title, key, color, icon) in enumerate(kpi_config):
            card = CTkFrame(kpi_frame, fg_color=COLORS["bg_white"],
                            corner_radius=8, border_width=1,
                            border_color=COLORS["border"])
            card.grid(row=0, column=i, padx=4, sticky="nsew")

            accent = CTkFrame(card, fg_color=color, corner_radius=0, height=4)
            accent.pack(fill="x")

            CTkLabel(card, text=f"{icon}  {title}",
                     font=CTkFont(family=FONT_FAMILY, size=11, weight="bold"),
                     text_color=COLORS["text_medium"]).pack(pady=(10, 2))

            val_label = CTkLabel(card, text="–",
                                 font=CTkFont(family=FONT_FAMILY, size=22, weight="bold"),
                                 text_color=color)
            val_label.pack(pady=(0, 12))
            self.kpi_labels[key] = val_label

        # ── Table ────────────────────────────────────────────
        table_outer = CTkFrame(main_frame, fg_color=COLORS["bg_white"],
                               corner_radius=8, border_width=1,
                               border_color=COLORS["border"])
        table_outer.grid(row=2, column=0, sticky="nsew")
        table_outer.grid_columnconfigure(0, weight=1)
        table_outer.grid_rowconfigure(1, weight=1)

        title_bar = CTkFrame(table_outer, fg_color=COLORS["header_blue"],
                             corner_radius=0, height=36)
        title_bar.grid(row=0, column=0, sticky="ew")
        title_bar.grid_propagate(False)
        CTkLabel(title_bar, text="  รายละเอียดข้อมูล Cost Benchmark",
                 font=CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
                 text_color="#FFFFFF").pack(side="left", padx=10, pady=8)

        table_inner = tk.Frame(table_outer, bg=COLORS["bg_white"])
        table_inner.grid(row=1, column=0, sticky="nsew", padx=2, pady=2)

        if HAS_TKSHEET:
            self._build_tksheet(table_inner)
        else:
            tk.Label(table_inner, text="⚠️ กรุณาติดตั้ง tksheet: pip install tksheet",
                     bg=COLORS["bg_white"], fg="red",
                     font=(FONT_FAMILY, 12)).pack(expand=True, pady=40)

    # =========================================================
    # TKSHEET TABLE
    # =========================================================
    def _build_tksheet(self, parent):
        self.display_columns = [
            "Selec\nt",
            "วันที่ขอราคา",
            "PRIORI\nTY",
            "WIN RATE\n%",
            "สถานะ",
            "ชื่อ\nSale",
            "Order No.",
            "Sale Order No.",
            "รายการสินค้า",
            "จำนวน",
            "Max of\nน้ำหนัก\n/เส้น",
            "Average of\nราคาขาย /\nเส้น",
            "ราคาขาย /\nกก.",
            "Sum of\nราคาขาย\nรวม",
            "Total\nWin Sales", # 🟢 เพิ่มคอลัมน์นี้ตรงกลาง
            "Average of\nMarkup\nGuide (%)",
            "Sum of\nต้นทุนรวม\n(รวมย้าย)",
            "avg. cost/per\norder",          # 🟢 เพิ่มใหม่
            "Price\nCompetitive\nness",     # 🟢 เพิ่มใหม่
            "Sum of\nต้นทุน/เส้น",
            "ต้นทุน/\nกก.",    
            "ชื่อ Supplier",
            "Average of\nส่วนลด 2 (%)",
            "Average of\nส่วนลด 1 (%)",
            "Average of\nส่วนลด 1\n(บาท)",          # 🟢 เพิ่มใหม่
            "Average of\nส่วนลด 2\n(บาท)",          # 🟢 เพิ่มใหม่
            "ส่วนลดรวม\n1+2",                        # 🟢 calculated
            "ส่วนลดรวม\n1+2 (%)",                   # 🟢 calculated
            "Sum of ต้นทุนรวม\n(ไม่รวมย้าย)",
        ]

        self.col_source = {
            "Selec\nt":                        "Select",   # ← key ต้องตรงกับ display_columns
            "วันที่ขอราคา":                      "วันที่ขอราคา",
            "PRIORI\nTY":                        "PRIORITY",
            "WIN RATE\n%":                        "WIN RATE %",
            "สถานะ":                             "สถานะ",
            "ชื่อ\nSale":                       "ชื่อ Sale",
            "Order No.":                          "Order No.",
            "Sale Order No.":                     "Sale Order No.",
            "รายการสินค้า":                      "รายการสินค้า",
            "จำนวน":                             "จำนวน",
            "Max of\nน้ำหนัก\n/เส้น":           "น้ำหนัก/เส้น",
            "Average of\nราคาขาย /\nเส้น":      "ราคาขาย / เส้น",
            "ราคาขาย /\nกก.":                   "ราคาขาย / กก.",
            "Sum of\nราคาขาย\nรวม":             "ราคาขาย รวม",
            "Total\nWin Sales":                 "ราคาขาย รวม",
            "Average of\nMarkup\nGuide (%)":     "Markup Guide (%)",
            "Sum of\nต้นทุนรวม\n(รวมย้าย)":    "ต้นทุนรวม (รวมย้าย)",
            "avg. cost/per\norder":              None,   # 🟢 calculated
            "Price\nCompetitive\nness":         None,   # 🟢 calculated
            "Sum of\nต้นทุน/เส้น":              "ต้นทุน/เส้น",
            "ต้นทุน/\nกก.":                     "ต้นทุน/กก. (รวมย้าย)",
            "ชื่อ Supplier":                     "ชื่อ Supplier",
            "Average of\nส่วนลด 2 (%)":         "ส่วนลด 2 (%)",
            "Average of\nส่วนลด 1 (%)":         "ส่วนลด 1 (%)",
            "Average of\nส่วนลด 1\n(บาท)":      "ส่วนลด 1 (บาท)",       # 🟢 เพิ่มใหม่
            "Average of\nส่วนลด 2\n(บาท)":      "ส่วนลด 2 (บาท)",       # 🟢 เพิ่มใหม่
            "ส่วนลดรวม\n1+2":                   None,                    # 🟢 calculated
            "ส่วนลดรวม\n1+2 (%)":              None,                    # 🟢 calculated
            "Sum of ต้นทุนรวม\n(ไม่รวมย้าย)":  "ต้นทุนรวม (ไม่รวมย้าย)",
        }

        self.money_cols = {
            "Sum of\nราคาขาย\nรวม",
            "Total\nWin Sales",
            "Sum of\nต้นทุนรวม\n(รวมย้าย)",
            "Average of\nราคาขาย /\nเส้น",
            "ราคาขาย /\nกก.",
            "Sum of\nต้นทุน/เส้น",
            "ต้นทุน/\nกก.",
            "Sum of ต้นทุนรวม\n(ไม่รวมย้าย)",
            "Average of\nส่วนลด 1\n(บาท)",           # 🟢 เพิ่มใหม่
            "Average of\nส่วนลด 2\n(บาท)",           # 🟢 เพิ่มใหม่
            "ส่วนลดรวม\n1+2",                         # 🟢 เพิ่มใหม่
            "avg. cost/per\norder",                    # 🟢 เพิ่มใหม่
        }
        self.pct_cols = {
            "WIN RATE\n%",
            "Average of\nMarkup\nGuide (%)",
            "Average of\nส่วนลด 2 (%)",
            "Average of\nส่วนลด 1 (%)",
            "ส่วนลดรวม\n1+2 (%)",                    # 🟢 เพิ่มใหม่
            "Price\nCompetitive\nness",              # 🟢 เพิ่มใหม่
        }
        self.num_cols = {
            "จำนวน",
            "Max of\nน้ำหนัก\n/เส้น",
            "Average of\nราคาขาย /\nเส้น",
        }

        self.sheet = Sheet(
            parent,
            headers=self.display_columns,
            data=[],
            theme="light",
            font=(FONT_FAMILY, 10, "normal"),
            header_font=(FONT_FAMILY, 9, "bold"),
            show_row_index=True,
            row_index_width=32,
            empty_horizontal=0,
            empty_vertical=10,
            row_height=26,
            header_height=52,
            # ทำให้ selection row ไม่เด่น — ใช้สีขาว/เทาอ่อนแทนน้ำเงิน
            table_selected_rows_bg="#F1F5F9",
            table_selected_rows_fg=COLORS["text_dark"],
            table_selected_rows_border_fg=COLORS["border"],
            table_selected_cells_bg="#F1F5F9",
            table_selected_cells_fg=COLORS["text_dark"],
            table_selected_cells_border_fg=COLORS["border"],
        )
        self.sheet.pack(fill="both", expand=True)

        self.sheet.enable_bindings((
            "single_select", "drag_select", "multi_select",
            "row_select", "column_select", "column_width_resize",
            "arrowkeys", "right_click_popup_menu",
            "rc_select", "copy", "cut", "paste",
            "delete", "undo", "edit_cell",
            "column_drag_and_drop",
        ))

        def _copy_without_quotes(event=None):
            """Override tksheet copy — ไม่ใส่ CSV quotes"""
            try:
                selected = self.sheet.get_selected_cells()
                if not selected:
                    rows = self.sheet.get_selected_rows()
                    if not rows:
                        return
                    data = self.sheet.get_sheet_data()
                    lines = []
                    for r in sorted(rows):
                        if r < len(data):
                            lines.append("\t".join(str(v) for v in data[r]))
                    text = "\n".join(lines)
                else:
                    rows_data = {}
                    for (r, c) in selected:
                        rows_data.setdefault(r, {})[c] = self.sheet.get_cell_data(r, c)
                    lines = []
                    for r in sorted(rows_data):
                        cols = rows_data[r]
                        lines.append("\t".join(str(cols[c]) for c in sorted(cols)))
                    text = "\n".join(lines)
                self.sheet.clipboard_clear()
                self.sheet.clipboard_append(text)
            except Exception:
                pass

        self.sheet.extra_bindings([("copy", _copy_without_quotes)])

        self.sheet.set_options(
            header_bg=COLORS["header_blue"],
            header_fg=COLORS["header_text"],
            header_grid_fg="#3B6FE8",
            table_bg=COLORS["bg_white"],
            table_fg=COLORS["text_dark"],
            table_grid_fg=COLORS["border"],
            index_bg="#F1F5F9",
            index_fg=COLORS["text_medium"],
            index_grid_fg=COLORS["border"],
            index_border_fg=COLORS["border"],
        )

        col_widths = {
            "Selec\nt":                          42,
            "วันที่ขอราคา":                      88,
            "PRIORI\nTY":                        62,
            "WIN RATE\n%":                        72,
            "สถานะ":                             52,
            "ชื่อ\nSale":                       52,
            "Order No.":                          88,
            "Sale Order No.":                    100,
            "รายการสินค้า":                     260,
            "จำนวน":                             68,
            "Max of\nน้ำหนัก\n/เส้น":           80,
            "Average of\nราคาขาย /\nเส้น":      95,
            "ราคาขาย /\nกก.":                   80,
            "Sum of\nราคาขาย\nรวม":            110,
            "Total\nWin Sales":                110,
            "Average of\nMarkup\nGuide (%)":    100,
            "Sum of\nต้นทุนรวม\n(รวมย้าย)":   120,
            "avg. cost/per\norder":             115,    # 🟢 เพิ่มใหม่
            "Price\nCompetitive\nness":        100,    # 🟢 เพิ่มใหม่
            "Sum of\nต้นทุน/เส้น":              90,
            "ต้นทุน/\nกก.":                     85, 
            "ชื่อ Supplier":                    160,
            "Average of\nส่วนลด 2 (%)":        105,
            "Average of\nส่วนลด 1 (%)":        105,
            "Average of\nส่วนลด 1\n(บาท)":     105,    # 🟢 เพิ่มใหม่
            "Average of\nส่วนลด 2\n(บาท)":     105,    # 🟢 เพิ่มใหม่
            "ส่วนลดรวม\n1+2":                   110,    # 🟢 เพิ่มใหม่
            "ส่วนลดรวม\n1+2 (%)":              100,    # 🟢 เพิ่มใหม่
            "Sum of ต้นทุนรวม\n(ไม่รวมย้าย)": 140,
        }
        for i, col in enumerate(self.display_columns):
            self.sheet.column_width(i, col_widths.get(col, 100))

    # =========================================================
    # DATA LOADING
    # =========================================================
    def _load_data_from_db(self):
        conn = self.app_container.get_connection()
        try:
            query = "SELECT * FROM cost_benchmarks"
            df = pd.read_sql(query, conn)

            # ── Normalize column names ───────────────────────────────────────
            # PostgreSQL อาจคืนชื่อ "select" (lowercase) แทน "Select"
            # เพราะ SELECT เป็น reserved word ที่ต้องใช้ double-quote ตอนสร้าง
            # ทำให้ case อาจไม่ตรงกับที่ dashboard ต้องการ → normalize ทั้งหมด
            df.columns = ["Select" if c.lower() == "select" else c for c in df.columns]
            # ────────────────────────────────────────────────────────────────

            try:
                # =========================================================
                # 📌 กฎการเปลี่ยนชื่อและซ่อน (ใช้กฎเดียวกับหน้า Cost Benchmark)
                # =========================================================
                hide_list = {"s", "charita-ct", "mp", "p"}
                rename_map = {
                    "jiraporn": "JN-IN-JIRAPORN",
                    "sale center": "CT-Sale Center",
                    "piyawan": "AM-IN-PIYAWAN",
                    "ilada": "ID-IN-ILADA",
                    "vow-s": "TG-IN-Tharinya",
                    "bunnycee": "KR-IN-Kanokporn",
                    "lethai": "LT-FL-LETHAI",
                    "wachira": "WA-FL-WACHIRA",
                    "vow-p": "TG-IN-Tharinya"  # 🟢 รวม VOW-P เข้ากับ VOW-S (Tharinya) ไปเลย!
                }

                if "รหัส Sale" in df.columns:
                    def map_sale_name(raw_name):
                        if pd.isna(raw_name): return ""
                        raw_str = str(raw_name).strip()
                        lower_str = raw_str.lower()
                        
                        if lower_str in hide_list:
                            return None # ข้อมูลที่โดนซ่อน จะให้กลายเป็น None เพื่อลบทิ้งทีหลัง
                            
                        if lower_str in rename_map:
                            return rename_map[lower_str]
                            
                        return raw_str

                    # แปลงร่างข้อมูลในคอลัมน์ "ชื่อ Sale"
                    df["ชื่อ Sale"] = df["รหัส Sale"].apply(map_sale_name)
                    
                    # ลบทิ้งบรรทัดที่เป็น None (คือพวกที่อยู่ใน hide_list) ออกจาก Dashboard เลย
                    df = df.dropna(subset=["ชื่อ Sale"])
                else:
                    df["ชื่อ Sale"] = ""
                    
            except Exception as e:
                print(f"sales map error: {e}")
                df["ชื่อ Sale"] = df.get("รหัส Sale", "")

            # กรอง created_by ที่ไม่ต้องการออก (mp, p)
            _hidden_pu = {"mp", "p"}
            if "created_by" in df.columns:
                df = df[~df["created_by"].astype(str).str.strip().str.lower().isin(_hidden_pu)]

            if df.empty:
                self.raw_df = pd.DataFrame()
            else:
                def clean_numeric(series):
                    return (
                        series.astype(str)
                        .str.strip()
                        .str.replace(',', '', regex=False)
                        .str.replace('%', '', regex=False)
                        .str.replace('฿', '', regex=False)
                        .str.replace(' ', '', regex=False)
                        .replace({'': '0', 'None': '0', 'nan': '0', 'NaN': '0', 'none': '0'})
                        .pipe(lambda s: pd.to_numeric(s, errors='coerce'))
                        .fillna(0)
                    )

                numeric_cols = [
                    "จำนวน", "ทุนรวม", "ราคาขาย รวม",
                    "ต้นทุนรวม (รวมย้าย)", "ต้นทุนรวม (ไม่รวมย้าย)",
                    "น้ำหนัก/เส้น", "ราคาขาย / เส้น", "ราคาขาย / กก.",
                    "WIN RATE %", "Markup Guide (%)",
                    "ต้นทุน/เส้น", "ส่วนลด 1 (%)", "ส่วนลด 2 (%)",
                    "ส่วนลด 1 (บาท)", "ส่วนลด 2 (บาท)",   # 🟢 เพิ่มใหม่
                    "ต้นทุน/กก. (รวมย้าย)", "ค่าย้าย (ซื้อ)",
                ]
                for col in numeric_cols:
                    if col in df.columns:
                        df[col] = clean_numeric(df[col])

                # ── Fallback: คำนวณ ต้นทุนรวม (รวมย้าย / ไม่รวมย้าย)
                # สำหรับ rows เก่าที่ formula ไม่เคยวิ่ง → DB เก็บเป็น NULL → clean_numeric → 0
                # สูตรเดียวกับ _auto_calculate_sheet:
                #   ต้นทุนรวม (รวมย้าย)    = (ต้นทุน/เส้น - d1 - d2) × qty + ค่าย้าย(ซื้อ)
                #   ต้นทุนรวม (ไม่รวมย้าย) = (ต้นทุน/เส้น - d1 - d2) × qty
                try:
                    if {"ต้นทุน/เส้น", "จำนวน"}.issubset(df.columns):
                        _qty  = df["จำนวน"]
                        _cost = df["ต้นทุน/เส้น"]
                        _d1   = df["ส่วนลด 1 (บาท)"] if "ส่วนลด 1 (บาท)" in df.columns else 0
                        _d2   = df["ส่วนลด 2 (บาท)"] if "ส่วนลด 2 (บาท)" in df.columns else 0
                        _move = df["ค่าย้าย (ซื้อ)"]  if "ค่าย้าย (ซื้อ)"  in df.columns else 0
                        _cost_after_disc = _cost - _d1 - _d2

                        if "ต้นทุนรวม (รวมย้าย)" in df.columns:
                            _computed_w = _cost_after_disc * _qty + _move
                            _fix_w = (df["ต้นทุนรวม (รวมย้าย)"] == 0) & (_computed_w > 0)
                            df.loc[_fix_w, "ต้นทุนรวม (รวมย้าย)"] = _computed_w[_fix_w]

                        if "ต้นทุนรวม (ไม่รวมย้าย)" in df.columns:
                            _computed_n = _cost_after_disc * _qty
                            _fix_n = (df["ต้นทุนรวม (ไม่รวมย้าย)"] == 0) & (_computed_n > 0)
                            df.loc[_fix_n, "ต้นทุนรวม (ไม่รวมย้าย)"] = _computed_n[_fix_n]
                except Exception as _fe:
                    print(f"[dashboard_cost fallback calc] {_fe}")

                self.raw_df = df

            self._update_filter_dropdowns()
            self._apply_filters()

        except Exception as e:
            messagebox.showerror("Database Error", f"โหลดข้อมูลล้มเหลว:\n{e}", parent=self)
        finally:
            if conn:
                self.app_container.release_connection(conn)

    # =========================================================
    # FILTER DROPDOWNS
    # =========================================================
    def _update_filter_dropdowns(self):
        if self.raw_df.empty:
            for key in self.filter_menus:
                self.filter_menus[key].configure(values=["All"])
                self.filter_vars[key].set("All")
            return

        def uniq(col):
            if col in self.raw_df.columns:
                # normalize ตัวพิมพ์ก่อน deduplicate
                vals = list(dict.fromkeys(
                    str(x).strip()
                    for x in self.raw_df[col]
                    if str(x).strip() not in ("", "None", "nan")
                ))
                
                # =========================================================
                # 📌 ถ้าเป็นคอลัมน์ "ชื่อ Sale" ให้จัดเรียงลำดับความสำคัญก่อน
                # =========================================================
                if col == "ชื่อ Sale":
                    def get_sort_key(name):
                        name_upper = str(name).upper()
                        if "-IN-" in name_upper: return 1      # Inbound มาก่อน
                        if "CT-" in name_upper or "-CT" in name_upper: return 2  # ตามด้วย Center
                        if name_upper == "STOCK": return 3     # ตามด้วย Stock
                        if "-FL-" in name_upper: return 4      # ตามด้วย Freelance
                        return 5                               # อื่นๆ
                    
                    return ["All"] + sorted(vals, key=lambda x: (get_sort_key(x), x.lower()))
                else:
                    return ["All"] + sorted(vals, key=lambda x: x.lower())
            return ["All"]

        # 🟢 ผูกชื่อตัวแปร Filter กับชื่อคอลัมน์ในตาราง Database
        mapping = {
            "order_no":      "Order No.",
            "sale_order_no": "Sale Order No.", 
            "product_name":  "รายการสินค้า",
            "pu_user":       "created_by",      
            "sale_name":     "ชื่อ Sale",
            "supplier":      "ชื่อ Supplier",
            "status":        "สถานะ",
            "priority":      "PRIORITY",
            "select":        "Select",
            "year":          "benchmark_year",
            "month":         "benchmark_month",
        }
        
        for key, col in mapping.items():
            vals = uniq(col)
            
            # เก็บ Data ทั้งหมดแยกไว้ สำหรับระบบพิมพ์ค้นหา
            if key == "order_no":
                self.all_order_nos = [v for v in vals if v != "All"]
            elif key == "sale_order_no": 
                self.all_sale_order_nos = [v for v in vals if v != "All"]
            elif key == "product_name": 
                self.all_product_names = [v for v in vals if v != "All"]
            elif key == "supplier": # 🟢 เก็บรายชื่อ Supplier 
                self.all_suppliers = [v for v in vals if v != "All"]
                
            self.filter_menus[key].configure(values=vals)
            
            # ถ้าค่าเดิมที่ตั้งไว้ไม่มีในลิสต์ ให้ปรับกลับเป็น All
            if self.filter_vars[key].get() not in vals and self.filter_vars[key].get() != "All":
                if key == "order_no" and self.filter_vars[key].get() in self.all_order_nos:
                    pass 
                elif key == "sale_order_no" and self.filter_vars[key].get() in self.all_sale_order_nos:
                    pass 
                elif key == "product_name" and self.filter_vars[key].get() in self.all_product_names: 
                    pass
                elif key == "supplier" and self.filter_vars[key].get() in self.all_suppliers: # 🟢 ปล่อยผ่านของ Supplier
                    pass
                else:
                    self.filter_vars[key].set("All")

    # =========================================================
    # APPLY FILTERS
    # =========================================================
    def _apply_filters(self, *args):
        if self.raw_df.empty:
            self._update_kpis(pd.DataFrame())
            self._update_table(pd.DataFrame())
            return

        df = self.raw_df.copy()

        date_from_str = self.filter_vars["date_from"].get()
        date_to_str = self.filter_vars["date_to"].get()

        if date_from_str or date_to_str:
            # 🟢 ฟังก์ชันอัจฉริยะ: แปลงวันที่จาก DB ให้เป็น ค.ศ. (รองรับ พ.ศ. 2569, ปีแบบย่อ 69, และ ค.ศ. 2026)
            def parse_thai_date(val):
                if pd.isna(val) or str(val).strip() in ("", "None", "nan"):
                    return pd.NaT
                
                # ตัดเอาเฉพาะวันที่ (เผื่อมีเวลาติดมา) และเปลี่ยน - เป็น /
                d_str = str(val).split()[0].replace("-", "/") 
                parts = d_str.split("/")
                
                if len(parts) == 3:
                    try:
                        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                        if y < 100: y += 2500        # ถ้าเป็นเลขย่อ เช่น 69 -> ให้บวกเป็น 2569
                        if y > 2400: y -= 543        # ถ้าเป็น พ.ศ. (เกิน 2400) -> ลบ 543 ให้กลายเป็น ค.ศ.
                        return pd.to_datetime(f"{y}-{m:02d}-{d:02d}", errors='coerce')
                    except ValueError:
                        return pd.to_datetime(val, errors='coerce')
                
                return pd.to_datetime(val, errors='coerce')

            try:
                # 1. ตรวจสอบเงื่อนไขวันที่ "ตั้งแต่ (From)"
                if date_from_str:
                    temp_dates_from = df["วันที่ขอราคา"].apply(parse_thai_date)
                    dt_from = pd.to_datetime(date_from_str, format='%d/%m/%Y')
                    df = df[temp_dates_from >= dt_from]
                    
                # 2. ตรวจสอบเงื่อนไขวันที่ "ถึง (To)"
                if date_to_str:
                    temp_dates_to = df["วันที่ขอราคา"].apply(parse_thai_date)
                    dt_to = pd.to_datetime(date_to_str, format='%d/%m/%Y')
                    # บวกเวลาให้ dt_to ไปจบที่ 23:59:59 ของวันนั้น (คลุมข้อมูลให้ครบถ้วน)
                    dt_to = dt_to.replace(hour=23, minute=59, second=59)
                    df = df[temp_dates_to <= dt_to]
                    
            except Exception as e:
                print(f"Date parsing error: {e}")

        # 🟢 กรองข้อมูลตาม Dropdown ที่เลือก
        col_map = {
            "order_no":      "Order No.",
            "sale_order_no": "Sale Order No.", 
            "product_name":  "รายการสินค้า",
            "pu_user":       "created_by",
            "sale_name":     "ชื่อ Sale",
            "supplier":      "ชื่อ Supplier",
            "status":        "สถานะ",
            "priority":      "PRIORITY",
            "select":        "Select",
            "year":          "benchmark_year",
            "month":         "benchmark_month",
        }
        
        for key, col in col_map.items():
            val = self.filter_vars[key].get()
            if val != "All" and val != "" and col in df.columns:
                # 🟢 เพิ่ม supplier ให้ค้นหาข้อความบางส่วนได้ด้วย
                if key in ["order_no", "sale_order_no", "product_name", "supplier"]: 
                    df = df[df[col].astype(str).str.contains(val, case=False, na=False)]
                else:
                    df = df[df[col].astype(str) == str(val)]

        self.current_filtered_df = df.copy()   # เก็บไว้สำหรับ Export
        self._update_kpis(df)
        self._update_table(df)

    # =========================================================
    # KPI UPDATE
    # =========================================================
    def _update_kpis(self, df):
        if df.empty:
            for key in self.kpi_labels:
                self.kpi_labels[key].configure(text="–")
            return

        df_active = df
        if "รายการสินค้า" in df.columns:
            df_active = df[df["รายการสินค้า"].astype(str).str.strip() != ""]

        # นับ Unique Sale Order No. (ไม่นับซ้ำ) — กัน "nan"/"None" จาก pandas astype(str)
        if "Sale Order No." in df_active.columns:
            _so = df_active["Sale Order No."].astype(str).str.strip()
            total_orders = _so[~_so.isin(["", "nan", "None", "NaN", "NaT"])].nunique()
        else:
            total_orders = len(df_active)

        total_qty    = (df_active["จำนวน"].sum()
                        if "จำนวน" in df_active.columns else 0)
        total_cost   = (df_active["ต้นทุนรวม (รวมย้าย)"].sum()
                        if "ต้นทุนรวม (รวมย้าย)" in df_active.columns else 0)
        total_sales  = (df_active["ราคาขาย รวม"].sum()
                        if "ราคาขาย รวม" in df_active.columns else 0)
                        
        # 🟢 [เพิ่มใหม่] คำนวณหายอด Win รวม (กรองเฉพาะสถานะ WIN)
        if "สถานะ" in df_active.columns and "ราคาขาย รวม" in df_active.columns:
            win_df = df_active[df_active["สถานะ"].astype(str).str.strip().str.upper() == "WIN"]
            total_win = pd.to_numeric(win_df["ราคาขาย รวม"], errors='coerce').fillna(0).sum()
        else:
            total_win = 0
                        
        # คำนวณ Markup เฉลี่ย โดยไม่เอาเลข 0 มาคิด
        if "Markup Guide (%)" in df_active.columns:
            margin_series = pd.to_numeric(df_active["Markup Guide (%)"], errors='coerce').dropna()
            margin_non_zero = margin_series[margin_series != 0] # ตัดเลข 0 ทิ้ง
            avg_margin = margin_non_zero.mean() if not margin_non_zero.empty else 0
        else:
            avg_margin = 0

        self.kpi_labels["total_orders"].configure(text=f"{total_orders:,}")
        self.kpi_labels["total_qty"].configure(text=f"{total_qty:,.0f}")
        
        self.kpi_labels["total_cost"].configure(
            text=f"฿{total_cost/1_000_000:.2f}M" if total_cost >= 1_000_000 else f"฿{total_cost:,.0f}")
            
        self.kpi_labels["total_sales"].configure(
            text=f"฿{total_sales/1_000_000:.2f}M" if total_sales >= 1_000_000 else f"฿{total_sales:,.0f}")
            
        # 🟢 [เพิ่มใหม่] จัดรูปแบบตัวเลขให้การ์ด ยอด Win รวม
        self.kpi_labels["total_win"].configure(
            text=f"฿{total_win/1_000_000:.2f}M" if total_win >= 1_000_000 else f"฿{total_win:,.0f}")
            
        self.kpi_labels["avg_margin"].configure(
            text=f"{avg_margin:,.2f}%" if pd.notna(avg_margin) else "0.00%")
    # =========================================================
    # TABLE UPDATE
    # =========================================================
    def _update_table(self, df):
        if not HAS_TKSHEET:
            return

        self.sheet.set_sheet_data([])
        self.sheet.deselect("all")   # ← ล้าง selection เก่าก่อนโหลดข้อมูลใหม่

        if df.empty:
            self.sheet.redraw()
            return

        col_source = self.col_source
        money_cols = self.money_cols
        pct_cols   = self.pct_cols
        num_cols   = self.num_cols

        # 🟢 pre-compute Price competitiveness per (SKU, จำนวน, Select type)
        # avg  = ค่าเฉลี่ยของทุก ✔ / เทียบเพื่อชุบ ✔ row ที่ (SKU, จำนวน, type) เดียวกัน
        # comp_avg_map = {(sku, qty, sel_type): avg_cost}
        SELECT_TYPES     = {"✔", "เทียบเพื่อชุบ ✔"}
        CHUB_TYPES       = {"เทียบเพื่อชุบ", "เทียบเพื่อชุบ ✔"}   # กลุ่มชุบทุก variant
        # {(sku, group): avg_cost}  group = "chub" | "normal"
        sku_avg_map  = {}
        comp_avg_map = {}
        if ("รายการสินค้า" in df.columns and "ต้นทุนรวม (รวมย้าย)" in df.columns
                and "Select" in df.columns):
            df_valid = df[df["รายการสินค้า"].astype(str).str.strip() != ""].copy()
            df_valid["ต้นทุนรวม (รวมย้าย)"] = pd.to_numeric(
                df_valid["ต้นทุนรวม (รวมย้าย)"], errors='coerce')
            df_valid["จำนวน"] = pd.to_numeric(df_valid.get("จำนวน", 0), errors='coerce')
            df_cost = df_valid[df_valid["ต้นทุนรวม (รวมย้าย)"] > 0].copy()
            df_cost["_grp"] = df_cost["Select"].astype(str).str.strip().apply(
                lambda s: "chub" if s in CHUB_TYPES else "normal")

            # avg แยกตาม (SKU, กลุ่ม) — ไม่ปน chub กับ normal
            grp_avg   = df_cost.groupby(["รายการสินค้า", "_grp"])["ต้นทุนรวม (รวมย้าย)"].mean()
            grp_count = df_cost.groupby(["รายการสินค้า", "_grp"])["ต้นทุนรวม (รวมย้าย)"].count()
            for (sku, grp), avg in grp_avg.items():
                sku_avg_map[(sku, grp)] = (avg, int(grp_count.get((sku, grp), 1)))

            # comp_avg_map ไม่ใช้งานแล้ว (เก็บไว้เผื่อ)
            comp_avg_map = {}

        # pre-compute: (SKU, grp) → row index ที่ถูก ✔ (ใช้แสดง avg. cost/per order)
        _sku_selected_idx = {}   # key = (sku, grp)
        for _idx, _r in df.iterrows():
            _sel = str(_r.get("Select", "")).strip()
            if _sel in SELECT_TYPES:
                _sku = str(_r.get("รายการสินค้า", "")).strip()
                _grp = "chub" if _sel in CHUB_TYPES else "normal"
                _key = (_sku, _grp)
                if _sku and _key not in _sku_selected_idx:
                    _sku_selected_idx[_key] = _idx

        table_data = []
        _shown_avg_sku = set()
        for _cur_idx, row in df.iterrows():
            row_data = []
            for dcol in self.display_columns:

                # 🟢 ส่วนลดรวม 1+2 = (ส่วนลด 1 บาท + ส่วนลด 2 บาท) × จำนวน
                if dcol == "ส่วนลดรวม\n1+2":
                    try:
                        d1  = pd.to_numeric(row.get("ส่วนลด 1 (บาท)", 0), errors='coerce') or 0
                        d2  = pd.to_numeric(row.get("ส่วนลด 2 (บาท)", 0), errors='coerce') or 0
                        qty = pd.to_numeric(row.get("จำนวน", 0), errors='coerce') or 0
                        val = (d1 + d2) * qty
                        row_data.append(f"฿{val:,.2f}" if val != 0 else "")
                    except Exception:
                        row_data.append("")
                    continue

                # 🟢 avg. cost/per order — แสดงที่ row ที่ถูก ✔ ของ (SKU, group) นั้น
                if dcol == "avg. cost/per\norder":
                    try:
                        sku     = str(row.get("รายการสินค้า", "")).strip()
                        sel_val = str(row.get("Select", "")).strip()
                        grp     = "chub" if sel_val in CHUB_TYPES else "normal"
                        map_key = (sku, grp)
                        entry   = sku_avg_map.get(map_key)
                        sel_idx = _sku_selected_idx.get(map_key)
                        # แสดงเฉพาะ row ที่ถูก ✔ (ถ้าไม่มี ✔ → แสดง row แรกของกลุ่ม)
                        if entry is not None and map_key not in _shown_avg_sku:
                            avg, cnt = entry
                            show_here = (sel_idx == _cur_idx) or (sel_idx is None)
                            if show_here:
                                _shown_avg_sku.add(map_key)
                                row_data.append(f"฿{avg:,.2f}")
                            else:
                                row_data.append("")
                        else:
                            row_data.append("")
                    except Exception:
                        row_data.append("")
                    continue

                # 🟢 Price competitiveness = (sel_cost - avg) / avg × 100
                # แสดงเฉพาะ row ที่เป็น ✔ หรือ เทียบเพื่อชุบ ✔
                # group key = (SKU, จำนวน, select type) — แยก ✔ vs เทียบเพื่อชุบ ✔
                if dcol == "Price\nCompetitive\nness":
                    try:
                        sel_type = str(row.get("Select", "")).strip()
                        if sel_type not in SELECT_TYPES:
                            row_data.append("")
                        else:
                            sku      = str(row.get("รายการสินค้า", "")).strip()
                            qty      = pd.to_numeric(row.get("จำนวน", 0), errors='coerce')
                            cost     = pd.to_numeric(row.get("ต้นทุนรวม (รวมย้าย)", 0), errors='coerce')
                            grp      = "chub" if sel_type in CHUB_TYPES else "normal"
                            entry    = sku_avg_map.get((sku, grp))
                            if entry and cost and not pd.isna(cost):
                                avg, cnt = entry
                                if avg != 0 and cnt > 1:
                                    pct = (cost - avg) / avg * 100
                                    row_data.append(f"{pct:+.2f}%" if pct != 0 else "0.00%")
                                else:
                                    row_data.append("")  # supplier เจ้าเดียว → ไม่มีตัวเปรียบ
                            else:
                                row_data.append("")
                    except Exception:
                        row_data.append("")
                    continue

                # 🟢 ส่วนลดรวม 1+2 (%) = ส่วนลดรวม / ทุนรวม × 100 (2 ทศนิยม)
                if dcol == "ส่วนลดรวม\n1+2 (%)":
                    try:
                        d1       = pd.to_numeric(row.get("ส่วนลด 1 (บาท)", 0), errors='coerce') or 0
                        d2       = pd.to_numeric(row.get("ส่วนลด 2 (บาท)", 0), errors='coerce') or 0
                        qty      = pd.to_numeric(row.get("จำนวน", 0), errors='coerce') or 0
                        tunkruam = pd.to_numeric(row.get("ทุนรวม", 0), errors='coerce') or 0
                        disc_sum = (d1 + d2) * qty
                        pct = (disc_sum / tunkruam * 100) if tunkruam != 0 else 0
                        row_data.append(f"{pct:.2f}%" if pct != 0 else "")
                    except Exception:
                        row_data.append("")
                    continue

                src = col_source.get(dcol)

                if src is None or src not in row.index:
                    row_data.append("")
                    continue

                val = row[src]

                # 🟢 เงื่อนไข Total Win Sales
                if dcol == "Total\nWin Sales":
                    status_val = str(row.get("สถานะ", "")).strip().upper()
                    if status_val != "WIN":
                        row_data.append("")
                        continue

                is_empty = pd.isna(val) or str(val).strip() in ("", "None", "nan", "NaN", "none")
                if is_empty:
                    row_data.append("")
                    continue

                if dcol in money_cols or dcol in pct_cols or dcol in num_cols:
                    fval = pd.to_numeric(val, errors='coerce')
                    if pd.isna(fval):
                        row_data.append(str(val))
                    elif dcol in money_cols:
                        # ฿0.00 หมายถึงคอลัมน์ formula ยังไม่ได้คำนวณ → แสดงเป็นว่าง
                        row_data.append(f"฿{float(fval):,.2f}" if fval != 0 else "")
                    elif dcol in pct_cols:
                        row_data.append(f"{float(fval):.2f}%")
                    else:
                        fv = float(fval)
                        row_data.append(f"{fv:,.2f}" if fv != int(fv) else f"{int(fv):,}")
                else:
                    row_data.append(str(val))

            if any(str(x).strip() for x in row_data):
                table_data.append(row_data)

        # ── Total row ──────────────────────────────────────────
        if table_data:
            def cidx(name):
                try: return self.display_columns.index(name)
                except ValueError: return -1

            total_row = [""] * len(self.display_columns)
            total_row[0] = "Total"

            sum_map = {
                "Sum of\nราคาขาย\nรวม":            "ราคาขาย รวม",
                "Total\nWin Sales":                "ราคาขาย รวม",
                "Sum of\nต้นทุนรวม\n(รวมย้าย)":   "ต้นทุนรวม (รวมย้าย)",
                "Sum of\nต้นทุน/เส้น":             "ต้นทุน/เส้น",
                "Sum of ต้นทุนรวม\n(ไม่รวมย้าย)": "ต้นทุนรวม (ไม่รวมย้าย)",
                "จำนวน":                            "จำนวน",
                "Max of\nน้ำหนัก\n/เส้น":          "น้ำหนัก/เส้น",
            }

            avg_map = {
                "Average of\nMarkup\nGuide (%)":   "Markup Guide (%)",
                "Average of\nส่วนลด 2 (%)":        "ส่วนลด 2 (%)",
                "Average of\nส่วนลด 1 (%)":        "ส่วนลด 1 (%)",
                "Average of\nส่วนลด 1\n(บาท)":     "ส่วนลด 1 (บาท)",   # 🟢 เพิ่มใหม่
                "Average of\nส่วนลด 2\n(บาท)":     "ส่วนลด 2 (บาท)",   # 🟢 เพิ่มใหม่
                "Average of\nราคาขาย /\nเส้น":     "ราคาขาย / เส้น",
                "ราคาขาย /\nกก.":                  "ราคาขาย / กก.",
            }

            # 🟢 SUM
            for dcol, src in sum_map.items():
                idx = cidx(dcol)
                if idx < 0 or src not in df.columns:
                    continue
                try:
                    if dcol == "Total\nWin Sales":
                        win_df = df[df["สถานะ"].astype(str).str.strip().str.upper() == "WIN"]
                        series = pd.to_numeric(win_df[src], errors='coerce').dropna()
                    else:
                        series = pd.to_numeric(df[src], errors='coerce').dropna()

                    if series.empty: continue
                    val = series.sum()
                    if val == 0: continue

                    if dcol in money_cols:
                        total_row[idx] = f"฿{float(val):,.2f}"
                    else:
                        fv = float(val)
                        total_row[idx] = f"{fv:,.2f}" if fv != int(fv) else f"{int(fv):,}"
                except Exception:
                    pass

            # 🟢 AVG (ไม่เอา 0) — ยกเว้น Markup ที่ต้องคำนวณแบบ weighted
            for dcol, src in avg_map.items():
                idx = cidx(dcol)
                if idx < 0 or src not in df.columns:
                    continue
                try:
                    # ── Average of Markup Guide — ใช้ Weighted Average ──────
                    # = (ราคาขายรวม - ต้นทุนรวม) / ต้นทุนรวม × 100
                    # ไม่ใช้ mean() ของแต่ละ row เพราะ row ที่มี qty ต่างกัน
                    # จะทำให้ค่าบิดเบือน (simple avg ≠ weighted avg)
                    if dcol == "Average of\nMarkup\nGuide (%)":
                        if "ราคาขาย รวม" in df.columns and "ต้นทุนรวม (รวมย้าย)" in df.columns:
                            total_s = pd.to_numeric(df["ราคาขาย รวม"], errors='coerce').fillna(0).sum()
                            total_c = pd.to_numeric(df["ต้นทุนรวม (รวมย้าย)"], errors='coerce').fillna(0).sum()
                            if total_c != 0:
                                markup_w = (total_s - total_c) / total_c * 100
                                total_row[idx] = f"{markup_w:.2f}%"
                        continue

                    series = pd.to_numeric(df[src], errors='coerce').dropna()
                    series_non_zero = series[series != 0]

                    if series_non_zero.empty:
                        continue

                    m = series_non_zero.mean()

                    if pd.notna(m):
                        if dcol in money_cols:
                            total_row[idx] = f"฿{float(m):,.2f}"
                        else:
                            total_row[idx] = f"{float(m):.2f}%"
                except Exception:
                    pass

            # 🟢 ส่วนลดรวม 1+2 รวม = sum of (ส่วนลด1+ส่วนลด2) × จำนวน
            try:
                idx_disc = cidx("ส่วนลดรวม\n1+2")
                if idx_disc >= 0 and "ส่วนลด 1 (บาท)" in df.columns and "จำนวน" in df.columns:
                    d1  = pd.to_numeric(df.get("ส่วนลด 1 (บาท)", 0), errors='coerce').fillna(0)
                    d2  = pd.to_numeric(df.get("ส่วนลด 2 (บาท)", 0), errors='coerce').fillna(0)
                    qty = pd.to_numeric(df.get("จำนวน", 0), errors='coerce').fillna(0)
                    total_disc = ((d1 + d2) * qty).sum()
                    if total_disc != 0:
                        total_row[idx_disc] = f"฿{total_disc:,.2f}"
            except Exception:
                pass

            # 🟢 ส่วนลดรวม 1+2 (%) รวม = total_disc / sum(ทุนรวม) × 100
            try:
                idx_pct = cidx("ส่วนลดรวม\n1+2 (%)")
                if idx_pct >= 0 and "ทุนรวม" in df.columns:
                    d1  = pd.to_numeric(df.get("ส่วนลด 1 (บาท)", 0), errors='coerce').fillna(0)
                    d2  = pd.to_numeric(df.get("ส่วนลด 2 (บาท)", 0), errors='coerce').fillna(0)
                    qty = pd.to_numeric(df.get("จำนวน", 0), errors='coerce').fillna(0)
                    tunkruam = pd.to_numeric(df["ทุนรวม"], errors='coerce').fillna(0)
                    total_disc = ((d1 + d2) * qty).sum()
                    total_tun  = tunkruam.sum()
                    if total_tun != 0:
                        pct = total_disc / total_tun * 100
                        total_row[idx_pct] = f"{pct:.2f}%"
            except Exception:
                pass

            # 🟢 cost/per order Total = avg จาก row ที่แสดงในตารางจริง (ไม่ใช่ raw df)
            try:
                idx_cpo = cidx("avg. cost/per\norder")
                if idx_cpo >= 0:
                    _cpo_vals = []
                    for _r in table_data:
                        _cell = str(_r[idx_cpo]).strip().replace("฿", "").replace(",", "")
                        try:
                            _v = float(_cell)
                            if _v > 0:
                                _cpo_vals.append(_v)
                        except Exception:
                            pass
                    if _cpo_vals:
                        _cpo_avg = sum(_cpo_vals) / len(_cpo_vals)
                        total_row[idx_cpo] = f"฿{_cpo_avg:,.2f}"
            except Exception:
                pass

            # 🟢 Price Competitive Total = weighted avg จาก row ที่แสดงในตารางจริง
            # = (Σ cost/per order × pct) / Σ cost/per order  ← weighted by cost
            try:
                idx_pc = cidx("Price\nCompetitive\nness")
                if idx_pc >= 0 and idx_cpo >= 0:
                    _pc_weighted_sum = 0.0
                    _pc_weight_total = 0.0
                    for _r in table_data:
                        _cost_cell = str(_r[idx_cpo]).strip().replace("฿", "").replace(",", "")
                        _pct_cell  = str(_r[idx_pc]).strip().replace("%", "")
                        try:
                            _cost = float(_cost_cell)
                            _pct  = float(_pct_cell)
                            if _cost > 0:
                                _pc_weighted_sum += _cost * _pct
                                _pc_weight_total += _cost
                        except Exception:
                            pass
                    if _pc_weight_total > 0:
                        _pc_total = _pc_weighted_sum / _pc_weight_total
                        total_row[idx_pc] = f"{_pc_total:.2f}%"
            except Exception:
                pass

            table_data.append(total_row)

        self.sheet.dehighlight_all()      # ล้าง highlight เก่าทั้งหมด (Total row สีเก่า ฯลฯ)
        self.sheet.deselect("all")        # ล้าง selection เก่า
        self.sheet.set_sheet_data(table_data)

        # ── Highlight columns ──────────────────────────────────
        blue_col  = "Sum of\nราคาขาย\nรวม"
        green_col = "Total\nWin Sales"
        red_col   = "Sum of\nต้นทุนรวม\n(รวมย้าย)"

        try:
            self.sheet.highlight_columns(
                columns=[self.display_columns.index(blue_col)],
                bg="#DBEAFE", fg="#1E40AF"
            )
        except ValueError:
            pass

        try:
            self.sheet.highlight_columns(
                columns=[self.display_columns.index(green_col)],
                bg="#DCFCE7", fg="#166534"
            )
        except ValueError:
            pass

        # 🔴 สีแดง (ต้นทุน)
        try:
            self.sheet.highlight_columns(
                columns=[self.display_columns.index(red_col)],
                bg=COLORS["highlight_red"],
                fg=COLORS["highlight_red_fg"],
            )
        except ValueError:
            pass

        # ── Alternate row shading ──────────────────────────────
        n_data = len(table_data) - 1

        skip_cols = {
            self.display_columns.index(blue_col) if blue_col in self.display_columns else -1,
            self.display_columns.index(green_col) if green_col in self.display_columns else -1,
            self.display_columns.index(red_col) if red_col in self.display_columns else -1,
        }

        for r in range(n_data):
            if r % 2 == 1:
                for c in range(len(self.display_columns)):
                    if c in skip_cols:
                        continue  # 👈 ข้ามคอลัมน์สี
                    self.sheet.highlight_cells(
                        row=r,
                        column=c,
                        bg="#F8FAFF",
                        fg=COLORS["text_dark"]
                    )

        # ── Total row highlight ────────────────────────────────
        total_row_idx = len(table_data) - 1
        
        # 1. ระบายสีพื้นฐานให้แถว Total ทั้งแถวเป็นสีน้ำเงินเข้ม (สำหรับคอลัมน์ทั่วไป)
        self.sheet.highlight_rows(
            rows=[total_row_idx],
            bg="#1E3A8A",          # พื้นน้ำเงินเข้ม
            fg="#FFFFFF",          # ตัวอักษรสีขาว
        )

        # 2. ไฮไลต์ทับ "เฉพาะช่อง Total" ของคอลัมน์พิเศษ ให้เป็นสีเดียวกับคอลัมน์แต่ "เข้มกว่า"
        try:
            blue_idx = self.display_columns.index(blue_col)
            self.sheet.highlight_cells(
                row=total_row_idx, 
                column=blue_idx, 
                bg="#3B82F6",      # สีฟ้าเข้ม (เข้มกว่า #DBEAFE ด้านบน)
                fg="#FFFFFF"       # อักษรสีขาวให้เห็นชัด
            )
        except ValueError:
            pass

        try:
            red_idx = self.display_columns.index(red_col)
            self.sheet.highlight_cells(
                row=total_row_idx, 
                column=red_idx, 
                bg="#EF4444",      # สีแดงเข้ม (เข้มกว่า #FEE2E2 ด้านบน)
                fg="#FFFFFF"       # อักษรสีขาว
            )
        except ValueError:
            pass
            
        try:
            # ช่อง Win Sales ปกติเป็นสีขาว แต่แถว Total จะให้เป็นเขียวเข้มเพื่อเน้น
            green_idx = self.display_columns.index(green_col)
            self.sheet.highlight_cells(
                row=total_row_idx, 
                column=green_idx, 
                bg="#10B981",      # สีเขียวเข้ม
                fg="#FFFFFF"       # อักษรสีขาว
            )
        except ValueError:
            pass

        self.sheet.readonly_columns(
            columns=list(range(len(self.display_columns))),
            readonly=True,
        )

        # ── ล้าง row selection ที่ค้างมาจากครั้งก่อน ──────────────
        # tksheet ธีม "light blue" จะระบายสีน้ำเงินทั้งแถวเมื่อ user เคยคลิก
        # และ selection นั้นจะคงอยู่แม้ refresh ข้อมูลแล้ว → deselect ทุกครั้ง
        self.sheet.deselect("all")

        self.sheet.redraw()